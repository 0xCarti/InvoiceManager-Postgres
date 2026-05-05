import csv
import datetime
from dataclasses import dataclass
from pathlib import Path
from typing import IO, Iterable, List, Optional

from app import db
from werkzeug.datastructures import FileStorage

from app.models import Item, Vendor, VendorItemAlias
from app.utils.pos_import import normalize_pos_alias
from app.utils.numeric import coerce_float


@dataclass
class ParsedPurchaseLine:
    vendor_sku: Optional[str]
    vendor_description: str
    pack_size: Optional[str]
    quantity: float
    unit_cost: Optional[float] = None


@dataclass
class ResolvedPurchaseLine:
    parsed_line: ParsedPurchaseLine
    alias: Optional[VendorItemAlias]
    item_id: Optional[int] = None
    unit_id: Optional[int] = None
    cost: Optional[float] = None


@dataclass
class ParsedPurchaseOrder:
    items: List[ParsedPurchaseLine]
    order_date: Optional[datetime.date] = None
    expected_date: Optional[datetime.date] = None
    order_number: Optional[str] = None
    expected_total: Optional[float] = None


class CSVImportError(Exception):
    """Raised when a CSV import cannot be processed."""


PURCHASE_IMPORT_PROFILE_SYSCO_SOURCE = "sysco_source"
PURCHASE_IMPORT_PROFILE_SYSCO_SHOP = "sysco_shop"
PURCHASE_IMPORT_PROFILE_PRATTS = "pratts"
PURCHASE_IMPORT_PROFILE_MANITOBA_LIQUOR = "manitoba_liquor"

PURCHASE_IMPORT_PROFILE_LABELS = {
    PURCHASE_IMPORT_PROFILE_SYSCO_SOURCE: "Sysco Source",
    PURCHASE_IMPORT_PROFILE_SYSCO_SHOP: "Sysco Shop",
    PURCHASE_IMPORT_PROFILE_PRATTS: "Pratts",
    PURCHASE_IMPORT_PROFILE_MANITOBA_LIQUOR: "Manitoba Liquor & Lotteries",
}

_SYSCO_REQUIRED_HEADERS = {
    "item": {"item"},
    "description": {"description", "item description"},
    "quantity": {"qty ship", "qty shipped", "quantity shipped"},
    "price": {"price", "unit price"},
}

_SYSCO_OPTIONAL_HEADERS = {
    "ext_price": {"ext price", "extended price", "extd price"},
    "order_number": {
        "order #",
        "po number",
        "order number",
        "po #",
        "host order",
    },
}

_PRATTS_REQUIRED_HEADERS = {
    "item": {"item"},
    "pack": {"pack"},
    "size": {"size"},
    "brand": {"brand"},
    "description": {"description"},
    "quantity": {"qty ship", "qty shipped", "quantity shipped"},
    "price": {"price", "unit price"},
    "ext_price": {"ext price", "extended price", "extd price"},
}

_PRATTS_OPTIONAL_HEADERS = {
    "order_number": {
        "order #",
        "po number",
        "order number",
        "po #",
        "host order",
    },
}

_SYSCO_SHOP_REQUIRED_HEADERS = {
    "item": {"supc"},
    "case_qty": {"case qty"},
    "split_qty": {"split qty"},
    "pack_size": {"pack/size", "pack size"},
    "description": {"description"},
}

_SYSCO_SHOP_OPTIONAL_HEADERS = {
    "brand": {"brand"},
    "case_price": {"case $", "case price"},
    "each_price": {"each $", "each price"},
}

_MANITOBA_LIQUOR_REQUIRED_HEADERS = {
    "item": {"item number"},
    "description": {"product description"},
    "quantity": {"order quantity"},
    "price": {"unit price"},
    "ext_price": {"extended price"},
}

_MANITOBA_LIQUOR_OPTIONAL_HEADERS = {
    "pack_size": {
        "package size",
        "vol/case size",
        "case size",
        "vol / case size",
    },
    "order_number": {"order no.", "order no", "order number"},
    "original_order_number": {
        "original order no",
        "original order no.",
        "original order number",
    },
    "invoice_date": {"invoice date"},
}

_MANITOBA_LIQUOR_SUMMARY_DESCRIPTIONS = {
    "subtotal",
    "taxable amount",
    "invoice total",
}


def _prepare_reader(file_obj: IO) -> csv.DictReader:
    file_obj.seek(0)
    return csv.DictReader(
        (line.decode("utf-8", errors="ignore") if isinstance(line, bytes) else line
         for line in file_obj),
    )


def _normalize_header_name(header: str) -> str:
    return " ".join(header.lstrip("\ufeff").strip().lower().split())


def _normalize_headers(headers):
    return {_normalize_header_name(header): header for header in headers or []}


def _resolve_headers(header_map: dict, required: dict, optional: dict | None = None):
    resolved = {}
    missing = []

    for name, aliases in required.items():
        normalized_aliases = {_normalize_header_name(alias) for alias in aliases}
        match = next((header_map[alias] for alias in normalized_aliases if alias in header_map), None)
        if match:
            resolved[name] = match
        else:
            missing.append(name)

    if optional:
        for name, aliases in optional.items():
            normalized_aliases = {_normalize_header_name(alias) for alias in aliases}
            for alias in normalized_aliases:
                if alias in header_map:
                    resolved[name] = header_map[alias]
                    break

    return resolved, missing


def _upload_extension(file: FileStorage | None) -> str:
    filename = getattr(file, "filename", "") or ""
    return Path(filename).suffix.lower()


def _normalized_vendor_name(vendor: Vendor | None) -> str:
    if vendor is None:
        return ""
    return " ".join(filter(None, [vendor.first_name, vendor.last_name])).strip().lower()


def supported_purchase_import_profiles(vendor: Vendor | None) -> list[str]:
    vendor_name = _normalized_vendor_name(vendor)
    profiles: list[str] = []
    if "sysco" in vendor_name:
        profiles.extend(
            [
                PURCHASE_IMPORT_PROFILE_SYSCO_SOURCE,
                PURCHASE_IMPORT_PROFILE_SYSCO_SHOP,
            ]
        )
    if "pratt" in vendor_name:
        profiles.append(PURCHASE_IMPORT_PROFILE_PRATTS)
    if "mbll" in vendor_name or "manitoba liquor" in vendor_name or (
        "manitoba" in vendor_name and "lotter" in vendor_name
    ):
        profiles.append(PURCHASE_IMPORT_PROFILE_MANITOBA_LIQUOR)
    return profiles


def purchase_import_profile_label(profile: str) -> str:
    return PURCHASE_IMPORT_PROFILE_LABELS.get(
        profile, profile.replace("_", " ").title()
    )


def _iter_excel_rows(file_obj: IO, *, extension: str) -> Iterable[list[object]]:
    ext = (extension or "").lower()
    file_obj.seek(0)
    if ext == ".xls":
        try:
            import xlrd  # type: ignore
        except ModuleNotFoundError:
            try:
                from app.vendor import xlrd  # type: ignore
            except ImportError as exc:  # pragma: no cover - environment specific
                raise CSVImportError(
                    "Legacy Excel imports require xlrd. Install dependencies and try again."
                ) from exc

        try:
            workbook_bytes = file_obj.read()
            book = xlrd.open_workbook(file_contents=workbook_bytes)
        except Exception as exc:
            raise CSVImportError(
                "Could not read the Excel workbook. Please upload the standard Manitoba Liquor & Lotteries export."
            ) from exc

        try:
            try:
                sheet = book.sheet_by_index(0)
            except IndexError as exc:
                raise CSVImportError("The workbook is empty.") from exc

            for row_idx in range(sheet.nrows):
                row: list[object] = []
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell(row_idx, col_idx)
                    value = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            value = xlrd.xldate.xldate_as_datetime(
                                value, book.datemode
                            )
                        except Exception:
                            pass
                    row.append(value)
                yield row
        finally:  # pragma: no branch - defensive cleanup
            try:
                book.release_resources()
            except AttributeError:
                pass
        return

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment specific
        raise CSVImportError(
            "Excel imports require openpyxl. Install dependencies and try again."
        ) from exc

    try:
        workbook = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as exc:
        raise CSVImportError(
            "Could not read the Excel workbook. Please upload the standard Manitoba Liquor & Lotteries export."
        ) from exc

    try:
        sheet = workbook.active
        try:
            if sheet.calculate_dimension() == "A1:A1":
                # Manitoba's export reports the used range incorrectly.
                sheet.reset_dimensions()
        except Exception:
            pass

        for row in sheet.iter_rows(values_only=True):
            yield list(row)
    finally:
        workbook.close()


def _row_value(row: list[object], index: int) -> object | None:
    if index < 0 or index >= len(row):
        return None
    return row[index]


def _stringify_cell(value: object | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned or None
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, "f").rstrip("0").rstrip(".")
    if isinstance(value, int):
        return str(value)
    return str(value).strip() or None


def _coerce_excel_date(value: object | None) -> datetime.date | None:
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel
        except ImportError:
            return None
        try:
            parsed = from_excel(value)
        except Exception:
            return None
        if isinstance(parsed, datetime.datetime):
            return parsed.date()
        if isinstance(parsed, datetime.date):
            return parsed
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        for fmt in (
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%m/%d/%Y",
            "%d/%m/%Y",
            "%b %d %Y %I:%M %p",
            "%B %d %Y %I:%M %p",
        ):
            try:
                return datetime.datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue
    return None


def _csv_value(
    row: dict[str, str],
    header_lookup: dict[str, str],
    name: str,
) -> str | None:
    header_name = header_lookup.get(name)
    if not header_name:
        return None
    return row.get(header_name)


def _iter_csv_rows(file_obj: IO) -> Iterable[list[str]]:
    file_obj.seek(0)
    reader = csv.reader(
        (
            line.decode("utf-8", errors="ignore")
            if isinstance(line, bytes)
            else line
            for line in file_obj
        )
    )
    for row in reader:
        yield row


def _is_manitoba_liquor_summary_row(description: str | None) -> bool:
    if not description:
        return False
    normalized = " ".join(description.strip().lower().split())
    if normalized in _MANITOBA_LIQUOR_SUMMARY_DESCRIPTIONS:
        return True
    return normalized.startswith("@ tax rate")


def _parse_manitoba_liquor_rows(
    records: Iterable[dict[str, object | None]],
    *,
    empty_message: str,
) -> ParsedPurchaseOrder:
    items: List[ParsedPurchaseLine] = []
    order_number = None
    order_date = None
    expected_total = 0.0
    has_ext_price = False

    for record in records:
        description = _stringify_cell(record.get("description"))
        if not description:
            continue

        if not order_number:
            order_number = _stringify_cell(record.get("order_number")) or _stringify_cell(
                record.get("original_order_number")
            )

        if order_date is None:
            order_date = _coerce_excel_date(record.get("invoice_date"))

        if _is_manitoba_liquor_summary_row(description):
            continue

        ext_cost = coerce_float(record.get("ext_price"))
        if ext_cost is not None:
            expected_total += ext_cost
            has_ext_price = True

        if description.strip().lower() == "container deposit":
            continue

        quantity = coerce_float(record.get("quantity"))
        if quantity is None or quantity <= 0:
            continue

        vendor_sku = _stringify_cell(record.get("item"))
        unit_cost = coerce_float(record.get("price"))
        pack_size = _stringify_cell(record.get("pack_size"))

        items.append(
            ParsedPurchaseLine(
                vendor_sku=vendor_sku or None,
                vendor_description=description or vendor_sku or "",
                pack_size=pack_size,
                quantity=quantity,
                unit_cost=unit_cost,
            )
        )

    if not items:
        raise CSVImportError(empty_message)

    return ParsedPurchaseOrder(
        items=items,
        order_number=order_number,
        order_date=order_date,
        expected_total=expected_total if has_ext_price else None,
    )


def _parse_sysco_csv(file_obj: IO) -> ParsedPurchaseOrder:
    reader = _prepare_reader(file_obj)
    header_map = _normalize_headers(reader.fieldnames)
    header_lookup, missing_headers = _resolve_headers(
        header_map, _SYSCO_REQUIRED_HEADERS, _SYSCO_OPTIONAL_HEADERS
    )
    if missing_headers:
        readable = ", ".join(sorted(missing_headers))
        raise CSVImportError(
            f"Missing required Sysco columns: {readable}. Please upload the standard export file."
        )

    items: List[ParsedPurchaseLine] = []
    order_number = None
    expected_total = 0.0
    has_ext_price = False
    for row in reader:
        raw_description = row.get(header_lookup["description"], "").strip()
        raw_qty = row.get(header_lookup["quantity"], "")
        raw_price = row.get(header_lookup["price"], "")
        vendor_sku = row.get(header_lookup["item"], "").strip()

        quantity = coerce_float(raw_qty)
        if quantity is None or quantity <= 0:
            continue

        if not order_number and "order_number" in header_lookup:
            order_number = row.get(header_lookup["order_number"], "").strip() or None

        unit_cost = coerce_float(raw_price)
        if "ext_price" in header_lookup:
            raw_ext = row.get(header_lookup["ext_price"], "")
            ext_cost = coerce_float(raw_ext)
            if ext_cost is not None:
                expected_total += ext_cost
                has_ext_price = True

        items.append(
            ParsedPurchaseLine(
                vendor_sku=vendor_sku or None,
                vendor_description=raw_description or vendor_sku,
                pack_size=None,
                quantity=quantity,
                unit_cost=unit_cost,
            )
        )

    if not items:
        raise CSVImportError("No purchasable lines found in the CSV file.")

    return ParsedPurchaseOrder(
        items=items,
        order_number=order_number,
        expected_total=expected_total if has_ext_price else None,
    )


def _parse_sysco_shop_csv(file_obj: IO) -> ParsedPurchaseOrder:
    header_row: list[str] | None = None
    header_record: list[str] | None = None
    product_rows: list[list[str]] = []

    for row in _iter_csv_rows(file_obj):
        if not row:
            continue
        record_type = (row[0] or "").strip().upper()
        if record_type == "H":
            header_record = row
        elif record_type == "F":
            header_row = row[1:]
        elif record_type == "P":
            product_rows.append(row[1:])

    if header_row is None:
        raise CSVImportError(
            "Missing the Sysco Shop field header row. Please upload the Sysco Shop export file."
        )

    header_map = _normalize_headers(header_row)
    header_lookup, missing_headers = _resolve_headers(
        header_map, _SYSCO_SHOP_REQUIRED_HEADERS, _SYSCO_SHOP_OPTIONAL_HEADERS
    )
    if missing_headers:
        readable = ", ".join(sorted(missing_headers))
        raise CSVImportError(
            "Missing required Sysco Shop columns: "
            f"{readable}. Please upload the Sysco Shop export file."
        )

    header_indexes = {
        name: header_row.index(header_name)
        for name, header_name in header_lookup.items()
    }

    items: List[ParsedPurchaseLine] = []
    for row in product_rows:
        vendor_sku = _stringify_cell(_row_value(row, header_indexes["item"]))
        description = _stringify_cell(
            _row_value(row, header_indexes["description"])
        ) or vendor_sku
        brand = _stringify_cell(_row_value(row, header_indexes.get("brand", -1)))
        pack_size = _stringify_cell(
            _row_value(row, header_indexes.get("pack_size", -1))
        )
        display_description = description
        if brand and description and not description.upper().startswith(brand.upper()):
            display_description = f"{brand} {description}"

        case_qty = coerce_float(_row_value(row, header_indexes["case_qty"])) or 0.0
        split_qty = coerce_float(_row_value(row, header_indexes["split_qty"])) or 0.0
        case_price = coerce_float(
            _row_value(row, header_indexes.get("case_price", -1))
        )
        each_price = coerce_float(
            _row_value(row, header_indexes.get("each_price", -1))
        )

        if case_qty > 0:
            items.append(
                ParsedPurchaseLine(
                    vendor_sku=vendor_sku or None,
                    vendor_description=display_description or vendor_sku or "",
                    pack_size=pack_size,
                    quantity=case_qty,
                    unit_cost=case_price,
                )
            )

        if split_qty > 0:
            split_description = display_description or vendor_sku or ""
            if case_qty > 0:
                split_description = f"{split_description} (split)"
            items.append(
                ParsedPurchaseLine(
                    vendor_sku=vendor_sku or None,
                    vendor_description=split_description,
                    pack_size=pack_size,
                    quantity=split_qty,
                    unit_cost=each_price if each_price is not None else case_price,
                )
            )

    if not items:
        raise CSVImportError("No purchasable lines found in the CSV file.")

    order_date = None
    expected_date = None
    order_number = None
    expected_total = None
    if header_record:
        order_date = _coerce_excel_date(_row_value(header_record, 4))
        expected_date = _coerce_excel_date(_row_value(header_record, 5))
        order_number = _stringify_cell(_row_value(header_record, 9)) or _stringify_cell(
            _row_value(header_record, 10)
        )
        expected_total = coerce_float(_row_value(header_record, 11))

    return ParsedPurchaseOrder(
        items=items,
        order_date=order_date,
        expected_date=expected_date,
        order_number=order_number,
        expected_total=expected_total,
    )


def _parse_pratts_csv(file_obj: IO) -> ParsedPurchaseOrder:
    reader = _prepare_reader(file_obj)
    header_map = _normalize_headers(reader.fieldnames)
    header_lookup, missing_headers = _resolve_headers(
        header_map, _PRATTS_REQUIRED_HEADERS, _PRATTS_OPTIONAL_HEADERS
    )
    if missing_headers:
        readable = ", ".join(sorted(missing_headers))
        raise CSVImportError(
            f"Missing required Pratts columns: {readable}. Please upload the standard export file."
        )

    items: List[ParsedPurchaseLine] = []
    order_number = None
    expected_total = 0.0
    for row in reader:
        vendor_sku = row.get(header_lookup["item"], "").strip()
        raw_description = row.get(header_lookup["description"], "").strip()
        raw_qty = row.get(header_lookup["quantity"], "")
        raw_price = row.get(header_lookup["price"], "")
        pack = row.get(header_lookup["pack"], "").strip()
        size = row.get(header_lookup["size"], "").strip()

        quantity = coerce_float(raw_qty)
        if quantity is None or quantity <= 0:
            continue

        if not order_number and "order_number" in header_lookup:
            order_number = row.get(header_lookup["order_number"], "").strip() or None

        pack_size = " ".join(filter(None, [pack, size])) or None
        unit_cost = coerce_float(raw_price)
        raw_ext = row.get(header_lookup["ext_price"], "")
        ext_cost = coerce_float(raw_ext)
        if ext_cost is not None:
            expected_total += ext_cost

        items.append(
            ParsedPurchaseLine(
                vendor_sku=vendor_sku or None,
                vendor_description=raw_description or vendor_sku,
                pack_size=pack_size,
                quantity=quantity,
                unit_cost=unit_cost,
            )
        )

    if not items:
        raise CSVImportError("No purchasable lines found in the CSV file.")

    return ParsedPurchaseOrder(
        items=items,
        order_number=order_number,
        expected_total=expected_total,
    )


def _parse_manitoba_liquor_csv(file_obj: IO) -> ParsedPurchaseOrder:
    reader = _prepare_reader(file_obj)
    header_map = _normalize_headers(reader.fieldnames)
    header_lookup, missing_headers = _resolve_headers(
        header_map,
        _MANITOBA_LIQUOR_REQUIRED_HEADERS,
        _MANITOBA_LIQUOR_OPTIONAL_HEADERS,
    )
    if missing_headers:
        readable = ", ".join(sorted(missing_headers))
        raise CSVImportError(
            "Missing required Manitoba Liquor & Lotteries columns: "
            f"{readable}. Please upload the standard export file."
        )

    records = (
        {
            "item": _csv_value(row, header_lookup, "item"),
            "description": _csv_value(row, header_lookup, "description"),
            "quantity": _csv_value(row, header_lookup, "quantity"),
            "price": _csv_value(row, header_lookup, "price"),
            "ext_price": _csv_value(row, header_lookup, "ext_price"),
            "pack_size": _csv_value(row, header_lookup, "pack_size"),
            "order_number": _csv_value(row, header_lookup, "order_number"),
            "original_order_number": _csv_value(
                row, header_lookup, "original_order_number"
            ),
            "invoice_date": _csv_value(row, header_lookup, "invoice_date"),
        }
        for row in reader
    )
    return _parse_manitoba_liquor_rows(
        records,
        empty_message="No purchasable lines found in the CSV file.",
    )


def _parse_manitoba_liquor_excel(file_obj: IO, *, extension: str) -> ParsedPurchaseOrder:
    rows = _iter_excel_rows(file_obj, extension=extension)
    header_row = next(rows, None)
    if header_row is None:
        raise CSVImportError("The workbook is empty.")

    header_map = _normalize_headers(
        [_stringify_cell(cell) or "" for cell in header_row]
    )
    header_lookup, missing_headers = _resolve_headers(
        header_map,
        _MANITOBA_LIQUOR_REQUIRED_HEADERS,
        _MANITOBA_LIQUOR_OPTIONAL_HEADERS,
    )
    if missing_headers:
        readable = ", ".join(sorted(missing_headers))
        raise CSVImportError(
            "Missing required Manitoba Liquor & Lotteries columns: "
            f"{readable}. Please upload the standard export file."
        )

    header_indexes = {
        name: header_row.index(header_name)
        for name, header_name in header_lookup.items()
    }

    records = (
        {
            "item": _row_value(row, header_indexes["item"]),
            "description": _row_value(row, header_indexes["description"]),
            "quantity": _row_value(row, header_indexes["quantity"]),
            "price": _row_value(row, header_indexes["price"]),
            "ext_price": _row_value(row, header_indexes["ext_price"]),
            "pack_size": _row_value(row, header_indexes.get("pack_size", -1)),
            "order_number": _row_value(row, header_indexes.get("order_number", -1)),
            "original_order_number": _row_value(
                row, header_indexes.get("original_order_number", -1)
            ),
            "invoice_date": _row_value(row, header_indexes.get("invoice_date", -1)),
        }
        for row in rows
    )
    return _parse_manitoba_liquor_rows(
        records,
        empty_message="No purchasable lines found in the workbook.",
    )


def _alias_rank(
    alias: VendorItemAlias,
    *,
    item_unit_id: int | None = None,
):
    updated_at = alias.updated_at or alias.created_at or datetime.datetime.min
    return (
        1 if item_unit_id and alias.item_unit_id == item_unit_id else 0,
        1 if alias.normalized_description else 0,
        1 if alias.item_unit_id is None else 0,
        updated_at,
        alias.id or 0,
    )


def parse_purchase_order_csv(
    file: FileStorage,
    vendor: Vendor,
    *,
    import_profile: str | None = None,
) -> ParsedPurchaseOrder:
    """Parse a vendor purchase-order upload into a purchase order structure."""

    if not file:
        raise CSVImportError("No file was provided for upload.")

    supported_profiles = supported_purchase_import_profiles(vendor)
    selected_profile = import_profile or (
        supported_profiles[0] if supported_profiles else None
    )

    if selected_profile and selected_profile not in supported_profiles:
        raise CSVImportError(
            "The selected import format is not supported for this vendor."
        )

    if selected_profile == PURCHASE_IMPORT_PROFILE_SYSCO_SOURCE:
        return _parse_sysco_csv(file.stream)
    if selected_profile == PURCHASE_IMPORT_PROFILE_SYSCO_SHOP:
        return _parse_sysco_shop_csv(file.stream)
    if selected_profile == PURCHASE_IMPORT_PROFILE_PRATTS:
        return _parse_pratts_csv(file.stream)
    if selected_profile == PURCHASE_IMPORT_PROFILE_MANITOBA_LIQUOR:
        extension = _upload_extension(file)
        if extension == ".csv":
            return _parse_manitoba_liquor_csv(file.stream)
        return _parse_manitoba_liquor_excel(file.stream, extension=extension)

    raise CSVImportError("Purchase-order imports are not yet supported for this vendor.")


def _default_unit_for_item(item: Item, preferred_unit_id: int | None = None) -> int | None:
    if preferred_unit_id:
        for unit in item.units:
            if unit.id == preferred_unit_id:
                return preferred_unit_id
    for unit in item.units:
        if unit.receiving_default:
            return unit.id
    return item.units[0].id if item.units else None


def normalize_vendor_alias_text(value: str | None) -> str:
    return normalize_pos_alias(value or "")


def find_preferred_vendor_alias(
    *,
    vendor: Vendor | None,
    item_id: int | None,
    item_unit_id: int | None,
) -> VendorItemAlias | None:
    if vendor is None or not item_id:
        return None

    aliases = VendorItemAlias.query.filter_by(
        vendor_id=vendor.id, item_id=item_id
    ).all()
    if not aliases:
        return None

    return max(aliases, key=lambda alias: _alias_rank(alias, item_unit_id=item_unit_id))


def preferred_vendor_aliases_for_items(
    *,
    item_ids: Iterable[int],
    vendor: Vendor | None = None,
    vendor_id: int | None = None,
) -> dict[int, VendorItemAlias]:
    resolved_vendor_id = vendor.id if vendor is not None else vendor_id
    normalized_item_ids: list[int] = []
    seen_item_ids: set[int] = set()
    for item_id in item_ids:
        if item_id is None or not str(item_id).strip():
            continue
        try:
            normalized_item_id = int(item_id)
        except (TypeError, ValueError):
            continue
        if normalized_item_id in seen_item_ids:
            continue
        seen_item_ids.add(normalized_item_id)
        normalized_item_ids.append(normalized_item_id)
    normalized_item_ids.sort()
    if not resolved_vendor_id or not normalized_item_ids:
        return {}

    aliases = VendorItemAlias.query.filter(
        VendorItemAlias.vendor_id == resolved_vendor_id,
        VendorItemAlias.item_id.in_(normalized_item_ids),
    ).all()

    alias_map: dict[int, VendorItemAlias] = {}
    for alias in aliases:
        current = alias_map.get(alias.item_id)
        if current is None or _alias_rank(alias) > _alias_rank(current):
            alias_map[alias.item_id] = alias

    return alias_map


def update_or_create_vendor_alias(
    *,
    vendor: Vendor,
    item_id: int,
    item_unit_id: int | None,
    vendor_sku: str | None,
    vendor_description: str | None,
    pack_size: str | None,
    default_cost: float | None,
) -> VendorItemAlias:
    normalized_description = normalize_vendor_alias_text(vendor_description or vendor_sku)

    alias_by_sku = None
    if vendor_sku:
        alias_by_sku = VendorItemAlias.query.filter_by(
            vendor_id=vendor.id, vendor_sku=vendor_sku
        ).first()
    alias_by_description = None
    if normalized_description:
        alias_by_description = VendorItemAlias.query.filter_by(
            vendor_id=vendor.id, normalized_description=normalized_description
        ).first()

    alias = alias_by_sku
    if alias is not None:
        alias.vendor_sku = vendor_sku or None
        alias.vendor_description = vendor_description or vendor_sku or alias.vendor_description
        if alias_by_description is None or alias_by_description.id == alias.id:
            alias.normalized_description = normalized_description or None
        alias.pack_size = pack_size or None
        alias.item_id = item_id
        alias.item_unit_id = item_unit_id
        alias.default_cost = default_cost
        return alias

    if alias_by_description is not None:
        alias_by_description.vendor_description = (
            vendor_description or vendor_sku or alias_by_description.vendor_description
        )
        alias_by_description.normalized_description = normalized_description or None
        alias_by_description.pack_size = pack_size or None
        alias_by_description.item_id = item_id
        alias_by_description.item_unit_id = item_unit_id
        alias_by_description.default_cost = default_cost

        if (
            vendor_sku
            and alias_by_description.vendor_sku
            and alias_by_description.vendor_sku != vendor_sku
        ):
            alias = VendorItemAlias(vendor_id=vendor.id)
            alias.vendor_sku = vendor_sku
            alias.vendor_description = (
                vendor_description
                or alias_by_description.vendor_description
                or vendor_sku
            )
            alias.normalized_description = None
            alias.pack_size = pack_size or alias_by_description.pack_size
            alias.item_id = item_id
            alias.item_unit_id = item_unit_id
            alias.default_cost = default_cost
            db.session.add(alias)
            return alias

        alias_by_description.vendor_sku = vendor_sku or None
        return alias_by_description

    alias = VendorItemAlias(vendor_id=vendor.id)
    alias.vendor_sku = vendor_sku or None
    alias.vendor_description = vendor_description or vendor_sku
    alias.normalized_description = normalized_description or None
    alias.pack_size = pack_size or None
    alias.item_id = item_id
    alias.item_unit_id = item_unit_id
    alias.default_cost = default_cost

    return alias


def resolve_vendor_purchase_lines(
    vendor: Vendor, parsed_lines: List[ParsedPurchaseLine]
) -> List[ResolvedPurchaseLine]:
    if not parsed_lines:
        return []

    vendor_aliases = VendorItemAlias.query.filter_by(vendor_id=vendor.id).all()
    alias_by_sku = {alias.vendor_sku: alias for alias in vendor_aliases if alias.vendor_sku}
    alias_by_description = {
        alias.normalized_description: alias
        for alias in vendor_aliases
        if alias.normalized_description
    }

    resolved: List[ResolvedPurchaseLine] = []
    for parsed_line in parsed_lines:
        normalized_description = normalize_vendor_alias_text(
            parsed_line.vendor_description
        )
        alias = None
        if parsed_line.vendor_sku:
            alias = alias_by_sku.get(parsed_line.vendor_sku)
        if alias is None and normalized_description:
            alias = alias_by_description.get(normalized_description)

        item_id = None
        unit_id = None
        resolved_cost = parsed_line.unit_cost

        if alias and alias.item:
            item_id = alias.item_id
            unit_id = _default_unit_for_item(alias.item, alias.item_unit_id)
            if resolved_cost is None:
                resolved_cost = alias.default_cost

        resolved.append(
            ResolvedPurchaseLine(
                parsed_line=parsed_line,
                alias=alias,
                item_id=item_id,
                unit_id=unit_id,
                cost=resolved_cost,
            )
        )

    return resolved


def serialize_parsed_line(line: ParsedPurchaseLine) -> dict:
    return {
        "vendor_sku": line.vendor_sku,
        "vendor_description": line.vendor_description,
        "pack_size": line.pack_size,
        "quantity": line.quantity,
        "unit_cost": line.unit_cost,
    }
