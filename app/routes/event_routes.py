import csv
import io
import json
import math
import os
from decimal import Decimal, ROUND_HALF_UP
from collections import defaultdict
from datetime import datetime
from secrets import token_urlsafe
from types import SimpleNamespace

from flask import (
    Blueprint,
    abort,
    current_app,
    flash,
    jsonify,
    make_response,
    redirect,
    render_template,
    request,
    session,
    url_for,
)
from flask_login import current_user, login_required
from werkzeug.utils import secure_filename

from app import db
from app.forms import (
    CSRFOnlyForm,
    EVENT_TYPES,
    EventForm,
    EventLocationConfirmForm,
    EventLocationForm,
    EventLocationUndoConfirmForm,
    ProductWithRecipeForm,
    UpdateOpeningCountsForm,
    ScanCountForm,
    TerminalSalesUploadForm,
)
from app.models import (
    Event,
    EventLocation,
    EventLocationTerminalSalesSummary,
    EventStandSheetItem,
    GLCode,
    Item,
    Location,
    LocationStandItem,
    Product,
    ProductRecipeItem,
    TerminalSale,
    TerminalSaleProductAlias,
    TerminalSaleLocationAlias,
    TerminalSalesResolutionState,
)
from app.services.pdf import render_stand_sheet_pdf
from app.utils.activity import log_activity
from app.utils.filter_state import (
    filters_to_query_args,
    get_filter_defaults,
    normalize_filters,
)
from app.utils.menu_assignments import (
    get_authoritative_location_products,
    get_location_drift_recipe_item_ids,
)
from app.utils.numeric import coerce_float
from app.utils.pos_import import (
    combine_terminal_sales_totals,
    derive_terminal_sales_quantity,
    extract_terminal_sales_location,
    group_terminal_sales_rows,
    normalize_pos_alias,
    parse_terminal_sales_number,
    terminal_sales_cell_is_blank,
)
from app.utils.units import (
    DEFAULT_BASE_UNIT_CONVERSIONS,
    convert_cost_for_reporting,
    convert_quantity,
    convert_quantity_for_reporting,
    get_unit_label,
)
from app.utils.text import build_text_match_predicate, normalize_name_for_sorting
from app.utils.email import send_email
from itsdangerous import BadSignature, URLSafeSerializer
from sqlalchemy import func, or_
from sqlalchemy.orm import Session, selectinload

_STAND_SHEET_FIELDS = (
    "opening_count",
    "transferred_in",
    "transferred_out",
    "adjustments",
    "eaten",
    "spoiled",
    "closing_count",
)

_TERMINAL_SALES_CONFLICT_GUIDANCE = (
    "POS sales imports do not include timestamps, so one location can only be "
    "balanced against one event per day. If the same physical location reopens "
    "later that day, combine those sessions into one event in the app."
)


def _format_date_span(start_date, end_date):
    """Return a compact date range label."""

    if start_date == end_date:
        return start_date.isoformat()
    return f"{start_date.isoformat()} to {end_date.isoformat()}"


def _find_terminal_sales_event_location_conflicts(
    *,
    start_date,
    end_date,
    location_ids,
    exclude_event_id=None,
):
    """Return overlapping event/location assignments that POS sales cannot split."""

    unique_location_ids = sorted({loc_id for loc_id in location_ids or [] if loc_id})
    if (
        not unique_location_ids
        or start_date is None
        or end_date is None
        or start_date > end_date
    ):
        return []

    query = (
        db.session.query(Event, EventLocation.location_id, Location.name)
        .join(EventLocation, EventLocation.event_id == Event.id)
        .join(Location, Location.id == EventLocation.location_id)
        .filter(EventLocation.location_id.in_(unique_location_ids))
        .filter(Event.start_date <= end_date, Event.end_date >= start_date)
        .order_by(Location.name.asc(), Event.start_date.asc(), Event.name.asc())
    )
    if exclude_event_id is not None:
        query = query.filter(Event.id != exclude_event_id)

    conflicts = []
    for event_obj, location_id, location_name in query.all():
        overlap_start = max(start_date, event_obj.start_date)
        overlap_end = min(end_date, event_obj.end_date)
        conflicts.append(
            {
                "event_id": event_obj.id,
                "event_name": event_obj.name,
                "location_id": location_id,
                "location_name": location_name,
                "event_start_date": event_obj.start_date,
                "event_end_date": event_obj.end_date,
                "overlap_start_date": overlap_start,
                "overlap_end_date": overlap_end,
            }
        )
    return conflicts


def _terminal_sales_conflicts_for_event(event_obj: Event):
    """Return overlapping location assignments for the given event."""

    location_ids = [event_location.location_id for event_location in event_obj.locations]
    return _find_terminal_sales_event_location_conflicts(
        start_date=event_obj.start_date,
        end_date=event_obj.end_date,
        location_ids=location_ids,
        exclude_event_id=event_obj.id,
    )


def _group_terminal_sales_conflicts_by_location(conflicts):
    """Return event-location conflicts grouped by location ID."""

    grouped: dict[int, list[dict]] = {}
    for conflict in conflicts:
        grouped.setdefault(conflict["location_id"], []).append(conflict)
    return grouped


def _build_terminal_sales_conflict_message(conflicts):
    """Return a user-facing explanation for ambiguous POS assignment conflicts."""

    if not conflicts:
        return _TERMINAL_SALES_CONFLICT_GUIDANCE

    detail_parts = []
    for conflict in conflicts[:3]:
        detail_parts.append(
            '%s overlaps with event "%s" on %s'
            % (
                conflict["location_name"],
                conflict["event_name"],
                _format_date_span(
                    conflict["overlap_start_date"],
                    conflict["overlap_end_date"],
                ),
            )
        )

    if len(conflicts) > 3:
        detail_parts.append(f"{len(conflicts) - 3} more conflict(s)")

    return (
        "POS sales cannot be split across multiple events for the same location "
        "on the same day because imported sales do not include timestamps. "
        f"{'; '.join(detail_parts)}. Combine the events in the app or remove "
        "the overlapping location."
    )


def suggest_terminal_sales_location_mapping(
    open_locations: list[EventLocation],
    sales_summary: dict[str, dict],
) -> dict[int, str]:
    """Return default mapping suggestions using stored location aliases."""

    if not open_locations:
        return {}

    sales_location_names = list(sales_summary.keys())
    if not sales_location_names:
        return {el.id: "" for el in open_locations}

    normalized_lookup = {
        name: normalize_pos_alias(name) for name in sales_location_names
    }
    normalized_to_originals: dict[str, list[str]] = {}
    for original, normalized in normalized_lookup.items():
        if not normalized:
            continue
        normalized_to_originals.setdefault(normalized, []).append(original)

    lowercase_lookup: dict[str, str] = {}
    ambiguous_lowercase: set[str] = set()
    for original in sales_location_names:
        lowered = original.casefold()
        existing = lowercase_lookup.get(lowered)
        if existing is None:
            lowercase_lookup[lowered] = original
        elif existing != original:
            ambiguous_lowercase.add(lowered)
    normalized_values = [
        norm for norm in normalized_lookup.values() if norm
    ]
    alias_lookup: dict[str, TerminalSaleLocationAlias] = {}
    if normalized_values:
        alias_rows = (
            TerminalSaleLocationAlias.query.filter(
                TerminalSaleLocationAlias.normalized_name.in_(normalized_values)
            ).all()
        )
        alias_lookup = {alias.normalized_name: alias for alias in alias_rows}

    default_mapping: dict[int, str] = {}
    for el in open_locations:
        assigned_value = ""
        location_obj = el.location
        if location_obj:
            if location_obj.name in sales_summary:
                assigned_value = location_obj.name
            else:
                normalized_location = normalize_pos_alias(location_obj.name)
                lowered_location = location_obj.name.casefold()

                if (
                    lowered_location in lowercase_lookup
                    and lowered_location not in ambiguous_lowercase
                ):
                    assigned_value = lowercase_lookup[lowered_location]

                if not assigned_value and normalized_location:
                    direct_candidates = normalized_to_originals.get(
                        normalized_location, []
                    )
                    if len(direct_candidates) == 1:
                        assigned_value = direct_candidates[0]
                    else:
                        for candidate in direct_candidates:
                            if candidate.casefold() == lowered_location:
                                assigned_value = candidate
                                break

                if not assigned_value and normalized_location:
                    alias = alias_lookup.get(normalized_location)
                    if (
                        alias is not None
                        and alias.location_id == location_obj.id
                    ):
                        candidates = normalized_to_originals.get(
                            normalized_location
                        )
                        if candidates:
                            assigned_value = candidates[0]
                        else:
                            assigned_value = alias.source_name

                if not assigned_value:
                    for sales_name, normalized in normalized_lookup.items():
                        if not normalized:
                            continue
                        alias = alias_lookup.get(normalized)
                        if (
                            alias
                            and alias.location_id == location_obj.id
                        ):
                            assigned_value = sales_name
                            break
        default_mapping[el.id] = assigned_value
    return default_mapping


def _conversion_mapping():
    """Return the configured reporting-unit conversions."""

    configured = current_app.config.get("BASE_UNIT_CONVERSIONS") or {}
    merged = dict(DEFAULT_BASE_UNIT_CONVERSIONS)
    merged.update(configured)
    return merged


_CURRENCY_QUANTIZE = Decimal("0.01")


def _quantize_currency(value: Decimal) -> Decimal:
    """Return a currency amount rounded to two decimal places."""

    return value.quantize(_CURRENCY_QUANTIZE, rounding=ROUND_HALF_UP)


def _convert_value_for_reporting(value, base_unit, conversions):
    """Convert a stored base-unit value for presentation."""

    if value is None or not base_unit:
        return value
    try:
        converted, _ = convert_quantity_for_reporting(
            float(value), base_unit, conversions
        )
    except (TypeError, ValueError):
        return value
    return converted


def _build_sheet_values(sheet, base_unit, conversions):
    """Return reporting-unit stand sheet values for display."""

    values = {}
    for field in _STAND_SHEET_FIELDS:
        raw = getattr(sheet, field, None) if sheet else None
        values[field] = (
            _convert_value_for_reporting(raw, base_unit, conversions)
            if raw is not None
            else None
        )
    return SimpleNamespace(**values)


def _build_stand_item_entry(
    *,
    item,
    expected=0.0,
    sales=0.0,
    sheet=None,
    recv_unit=None,
    trans_unit=None,
    conversions=None,
):
    """Assemble a stand-sheet entry enriched with reporting metadata."""

    conversions = conversions or _conversion_mapping()
    base_unit = item.base_unit
    report_unit = conversions.get(base_unit, base_unit)
    report_label = get_unit_label(report_unit)
    expected_display = _convert_value_for_reporting(expected, base_unit, conversions)
    sales_display = _convert_value_for_reporting(sales, base_unit, conversions)
    if sales_display is None:
        sales_display = 0.0
    return {
        "item": item,
        "expected": expected_display,
        "expected_base": expected,
        "sales": sales_display,
        "sales_base": sales,
        "sheet": sheet,
        "sheet_values": _build_sheet_values(sheet, base_unit, conversions),
        "base_unit": base_unit,
        "report_unit": report_unit,
        "report_unit_label": report_label,
        "recv_unit": recv_unit,
        "trans_unit": trans_unit,
    }


def _calculate_confirmed_sales_summary(event: Event) -> SimpleNamespace | None:
    """Return aggregated terminal sales for confirmed locations within an event."""

    total_quantity = 0.0
    total_amount = 0.0
    has_confirmed = False

    for event_location in event.locations:
        if not event_location.confirmed:
            continue
        has_confirmed = True
        for sale in event_location.terminal_sales:
            quantity = float(sale.quantity or 0.0)
            product = sale.product
            price = float(getattr(product, "price", 0.0) or 0.0) if product else 0.0
            total_quantity += quantity
            total_amount += quantity * price

    if not has_confirmed:
        return None

    return SimpleNamespace(quantity=total_quantity, amount=total_amount)


def _fallback_item_price(item, allowed_product_ids: set[int] | None = None) -> float | None:
    """Estimate a per-unit price for an inventory item based on recipe products."""

    if item is None:
        return None

    prices: list[float] = []
    for recipe in item.recipe_items:
        product = recipe.product
        if product is None:
            continue
        if allowed_product_ids is not None and product.id not in allowed_product_ids:
            continue
        factor = recipe.unit.factor if recipe.unit else 1.0
        units_per_product = float(recipe.quantity or 0.0) * factor
        if units_per_product <= 0:
            continue
        price = float(getattr(product, "price", 0.0) or 0.0)
        prices.append(price / units_per_product)

    if prices:
        return sum(prices) / len(prices)

    return None


def _build_item_price_lookup(
    event_location: EventLocation, stand_items: list[dict]
) -> dict[int, float]:
    """Return a mapping of item IDs to price-per-unit estimates for a location."""

    usage_totals: dict[int, float] = defaultdict(float)
    revenue_totals: dict[int, float] = defaultdict(float)

    for sale in event_location.terminal_sales:
        product = sale.product
        if product is None:
            continue
        quantity = float(sale.quantity or 0.0)
        price = float(getattr(product, "price", 0.0) or 0.0)
        if quantity == 0:
            continue
        sale_revenue = quantity * price
        for recipe in product.recipe_items:
            if not recipe.countable or recipe.item_id is None:
                continue
            factor = recipe.unit.factor if recipe.unit else 1.0
            units_per_product = float(recipe.quantity or 0.0) * factor
            if units_per_product <= 0:
                continue
            item_units = quantity * units_per_product
            usage_totals[recipe.item_id] += item_units
            revenue_totals[recipe.item_id] += sale_revenue

    price_lookup: dict[int, float] = {}
    for item_id, units in usage_totals.items():
        if units > 0:
            price_lookup[item_id] = revenue_totals[item_id] / units

    location_obj = event_location.location
    allowed_product_ids: set[int] | None = None
    if location_obj is not None:
        allowed_product_ids = {
            product.id for product in get_authoritative_location_products(location_obj)
        }

    for entry in stand_items:
        item = entry.get("item")
        if item is None:
            continue
        item_id = item.id
        if item_id in price_lookup:
            continue
        fallback_price = _fallback_item_price(item, allowed_product_ids)
        if fallback_price is not None:
            price_lookup[item_id] = fallback_price

    return price_lookup


def _calculate_physical_vs_terminal_variance(event: Event) -> float | None:
    """Return the total dollar variance for confirmed locations in an event."""

    total_variance = 0.0
    any_confirmed = False
    has_priced_variance = False

    for event_location in event.locations:
        if not event_location.confirmed:
            continue
        any_confirmed = True
        _, stand_items = _get_stand_items(event_location.location_id, event_location.event_id)
        price_lookup = _build_item_price_lookup(event_location, stand_items)
        for entry in stand_items:
            sheet = entry.get("sheet")
            if sheet is None:
                continue
            price_per_unit = price_lookup.get(sheet.item_id)
            if price_per_unit is None:
                continue
            variance_units = (
                float(sheet.opening_count or 0.0)
                + float(sheet.transferred_in or 0.0)
                + float(sheet.adjustments or 0.0)
                - float(sheet.transferred_out or 0.0)
                - float(entry.get("sales_base") or 0.0)
                - float(sheet.eaten or 0.0)
                - float(sheet.spoiled or 0.0)
                - float(sheet.closing_count or 0.0)
            )
            if variance_units == 0:
                continue
            total_variance += variance_units * price_per_unit
            has_priced_variance = True

    if not any_confirmed:
        return None

    if not has_priced_variance:
        return 0.0

    return total_variance


def _sync_event_location_opening_counts(event_location: EventLocation) -> int:
    """Ensure stand sheet opening counts mirror the location inventory."""

    inventory_records = LocationStandItem.query.filter_by(
        location_id=event_location.location_id
    ).all()
    if not inventory_records:
        return 0

    existing_sheets = {
        sheet.item_id: sheet
        for sheet in EventStandSheetItem.query.filter_by(
            event_location_id=event_location.id
        )
    }

    updated = 0
    for record in inventory_records:
        sheet = existing_sheets.get(record.item_id)
        if sheet is None:
            sheet = EventStandSheetItem(
                event_location_id=event_location.id,
                item_id=record.item_id,
            )
            db.session.add(sheet)
        sheet.opening_count = float(record.expected_count or 0.0)
        updated += 1

    return updated


def _convert_report_value_to_base(value, base_unit, report_unit):
    """Convert a reporting-unit form value back into the base unit."""

    if value is None:
        return 0.0
    if not base_unit or not report_unit or base_unit == report_unit:
        return value
    try:
        return convert_quantity(value, report_unit, base_unit)
    except (TypeError, ValueError):
        return value


def _ensure_location_items(location_obj: Location, product_obj: Product) -> None:
    """Ensure a location has inventory records for the product's countable items."""

    if location_obj is None or product_obj is None:
        return

    for recipe_item in product_obj.recipe_items:
        if not recipe_item.countable:
            continue
        record = LocationStandItem.query.filter_by(
            location_id=location_obj.id, item_id=recipe_item.item_id
        ).first()
        if record is None:
            db.session.add(
                LocationStandItem(
                    location_id=location_obj.id,
                    item_id=recipe_item.item_id,
                    expected_count=0,
                    purchase_gl_code_id=recipe_item.item.purchase_gl_code_id,
                )
            )


def _normalize_variance_details(value):
    """Return a dict variance payload, decoding JSON strings when needed."""

    if not value:
        return None

    normalized = value
    if isinstance(value, str):
        try:
            normalized = json.loads(value)
        except (TypeError, ValueError):
            return None

    if not isinstance(normalized, dict):
        return None

    return normalized


def _derive_summary_totals_from_details(
    details: dict | None,
) -> tuple[float | None, float | None]:
    """Derive total quantity and amount values from variance details."""

    details = _normalize_variance_details(details)
    if not details:
        return (None, None)

    def _coerce(value):
        try:
            if value is None:
                return None
            return float(value)
        except (TypeError, ValueError):
            return None

    total_quantity = 0.0
    total_amount = 0.0
    have_quantity = False
    have_amount = False

    def _accumulate(entries: list[dict]) -> None:
        nonlocal total_quantity, total_amount, have_quantity, have_amount
        for entry in entries:
            quantity_value = _coerce(entry.get("quantity"))
            if quantity_value is not None:
                total_quantity += quantity_value
                have_quantity = True

            amount_value = _coerce(entry.get("file_amount"))
            if amount_value is None and quantity_value is not None:
                for price in entry.get("file_prices") or []:
                    price_value = _coerce(price)
                    if price_value is not None:
                        amount_value = quantity_value * price_value
                        break

            if amount_value is not None:
                total_amount += amount_value
                have_amount = True

    _accumulate(details.get("products") or [])
    _accumulate(details.get("unmapped_products") or [])

    return (
        total_quantity if have_quantity else None,
        total_amount if have_amount else None,
    )


def _apply_pending_sales(
    pending_sales: list[dict] | None,
    pending_totals: list[dict] | None = None,
    *,
    link_products_to_locations: bool = False,
) -> set[str]:
    """Persist uploaded terminal sales for the provided event locations."""

    updated_locations: set[str] = set()
    totals_map: dict[int, dict] = {}
    event_location_ids: set[int] = set()

    sanitized_sales: list[dict] = []
    for entry in pending_sales or []:
        if not isinstance(entry, dict):
            continue
        sanitized_entry = dict(entry)
        try:
            el_id = int(sanitized_entry.get("event_location_id"))
        except (TypeError, ValueError):
            sanitized_sales.append(sanitized_entry)
            continue
        sanitized_entry["event_location_id"] = el_id
        sanitized_sales.append(sanitized_entry)
        event_location_ids.add(el_id)

    if pending_totals:
        for entry in pending_totals:
            try:
                el_id = int(entry.get("event_location_id"))
            except (TypeError, ValueError):
                continue
            totals_map[el_id] = entry
            event_location_ids.add(el_id)

    if event_location_ids:
        (
            TerminalSale.query.filter(
                TerminalSale.event_location_id.in_(event_location_ids)
            ).delete(synchronize_session=False)
        )
        (
            EventLocationTerminalSalesSummary.query.filter(
                EventLocationTerminalSalesSummary.event_location_id.in_(
                    event_location_ids
                )
            ).delete(synchronize_session=False)
        )
        db.session.flush()

    def _sanitize_variance_details(value):
        normalized = _normalize_variance_details(value)
        if not normalized:
            return None

        def _sanitize_prices(values):
            return [
                coerce_float(price)
                for price in (values or [])
                if price is not None
            ]

        sanitized_products: list[dict] = []
        for entry in normalized.get("products", []):
            sanitized_products.append(
                {
                    "product_id": entry.get("product_id"),
                    "product_name": entry.get("product_name"),
                    "quantity": coerce_float(entry.get("quantity")),
                    "file_amount": coerce_float(entry.get("file_amount")),
                    "file_prices": _sanitize_prices(entry.get("file_prices")),
                    "app_price": coerce_float(entry.get("app_price")),
                    "sales_location": entry.get("sales_location"),
                }
            )

        sanitized_price_mismatches: list[dict] = []
        for entry in normalized.get("price_mismatches", []):
            sanitized_price_mismatches.append(
                {
                    "product_id": entry.get("product_id"),
                    "product_name": entry.get("product_name"),
                    "file_prices": _sanitize_prices(entry.get("file_prices")),
                    "app_price": coerce_float(entry.get("app_price")),
                    "sales_location": entry.get("sales_location"),
                }
            )

        sanitized_menu_issues: list[dict] = []
        for entry in normalized.get("menu_issues", []):
            sanitized_menu_issues.append(
                {
                    "product_id": entry.get("product_id"),
                    "product_name": entry.get("product_name"),
                    "menu_name": entry.get("menu_name"),
                    "sales_location": entry.get("sales_location"),
                }
            )

        sanitized_unmapped: list[dict] = []
        for entry in normalized.get("unmapped_products", []):
            sanitized_unmapped.append(
                {
                    "product_name": entry.get("product_name"),
                    "quantity": coerce_float(entry.get("quantity")),
                    "file_amount": coerce_float(entry.get("file_amount")),
                    "file_prices": _sanitize_prices(entry.get("file_prices")),
                    "sales_location": entry.get("sales_location"),
                }
            )

        sanitized = {
            "products": sanitized_products,
            "price_mismatches": sanitized_price_mismatches,
            "menu_issues": sanitized_menu_issues,
            "unmapped_products": sanitized_unmapped,
        }

        if not any(sanitized.values()):
            return None
        return sanitized

    for entry in sanitized_sales:
        event_location_id = entry.get("event_location_id")
        product_id = entry.get("product_id")
        quantity_value = entry.get("quantity", 0.0)
        if not event_location_id:
            continue
        event_location = db.session.get(EventLocation, event_location_id)
        if event_location is None:
            continue
        product = None
        if product_id:
            product = db.session.get(Product, product_id)
        source_name = entry.get("source_name")
        product_name = entry.get("product_name") or source_name
        if product is None and product_name:
            product = Product.query.filter(Product.name == product_name).first()
        if product is None and source_name and source_name != product_name:
            product = Product.query.filter(Product.name == source_name).first()
        if product is None:
            if not product_name:
                continue
            price_value = coerce_float(entry.get("product_price")) or 0.0
            product = Product(
                name=product_name,
                price=price_value,
                invoice_sale_price=price_value,
                cost=0.0,
            )
            db.session.add(product)
            db.session.flush()
        normalized_name = (entry.get("normalized_name") or "").strip()
        if normalized_name:
            alias = TerminalSaleProductAlias.query.filter_by(
                normalized_name=normalized_name
            ).first()
            if alias is None:
                alias = TerminalSaleProductAlias(
                    source_name=source_name or product_name,
                    normalized_name=normalized_name,
                    product=product,
                )
                db.session.add(alias)
            else:
                alias.source_name = source_name or product_name
                alias.product = product
        entry["product_id"] = product.id
        entry.setdefault("product_name", product.name)
        location_obj = event_location.location
        sale = TerminalSale.query.filter_by(
            event_location_id=event_location.id, product_id=product.id
        ).first()
        if sale:
            sale.quantity = quantity_value
        else:
            db.session.add(
                TerminalSale(
                    event_location_id=event_location.id,
                    product_id=product.id,
                    quantity=quantity_value,
                    sold_at=datetime.utcnow(),
                )
            )
        if (
            link_products_to_locations
            and location_obj is not None
            and location_obj.current_menu is None
            and product not in location_obj.products
        ):
            location_obj.products.append(product)
            _ensure_location_items(location_obj, product)
        if location_obj is not None and location_obj.name:
            updated_locations.add(location_obj.name)

    if totals_map:
        for el_id, data in totals_map.items():
            summary = EventLocationTerminalSalesSummary.query.filter_by(
                event_location_id=el_id
            ).first()
            if summary is None:
                summary = EventLocationTerminalSalesSummary(
                    event_location_id=el_id
                )
                db.session.add(summary)
            summary.source_location = data.get("source_location")
            summary.total_quantity = coerce_float(data.get("total_quantity"))
            total_amount_value = coerce_float(data.get("total_amount"))
            net_total_value = coerce_float(data.get("net_including_tax_total"))
            discount_value = coerce_float(data.get("discount_total"))
            if total_amount_value is None:
                total_amount_value = combine_terminal_sales_totals(
                    net_total_value, discount_value
                )
            summary.total_amount = total_amount_value
            summary.variance_details = _sanitize_variance_details(
                data.get("variance_details")
            )
            fallback_quantity, fallback_amount = _derive_summary_totals_from_details(
                summary.variance_details
            )
            if summary.total_quantity is None and fallback_quantity is not None:
                summary.total_quantity = fallback_quantity
            if summary.total_amount is None and fallback_amount is not None:
                summary.total_amount = fallback_amount
    return updated_locations


def _apply_resolution_actions(issue_state: dict) -> tuple[list[str], list[str]]:
    """Apply queued price and menu resolutions for terminal sales."""

    price_updates: list[str] = []
    menu_updates: list[str] = []

    for location_issue in issue_state.get("queue", []):
        event_location_id = location_issue.get("event_location_id")
        event_location = None
        if event_location_id:
            event_location = db.session.get(EventLocation, event_location_id)

        for issue in location_issue.get("price_issues", []):
            if issue.get("resolution") != "update":
                continue
            product_id = issue.get("product_id")
            new_price = issue.get("selected_price")
            if new_price is None:
                new_price = issue.get("terminal_price")
            if new_price is None:
                new_price = issue.get("target_price")
            if product_id is None or new_price is None:
                continue
            product = db.session.get(Product, product_id)
            if product is None:
                continue
            coerced_price = coerce_float(new_price)
            if coerced_price is None:
                continue
            product.price = coerced_price
            price_updates.append(product.name)

        if event_location is None:
            continue

        location_obj = event_location.location
        if location_obj is None:
            continue

        for issue in location_issue.get("menu_issues", []):
            if issue.get("resolution") != "add":
                continue
            product_id = issue.get("product_id")
            if product_id is None:
                continue
            product = db.session.get(Product, product_id)
            if product is None:
                continue
            if product not in location_obj.products:
                location_obj.products.append(product)
            if (
                location_obj.current_menu
                and product not in location_obj.current_menu.products
            ):
                location_obj.current_menu.products.append(product)
            _ensure_location_items(location_obj, product)
            menu_updates.append(f"{product.name} @ {location_obj.name}")

    return price_updates, menu_updates


def _should_store_terminal_summary(
    loc_sales: dict | None,
    location_updated: bool,
    unmatched_entries: list[dict],
) -> bool:
    if location_updated:
        return True

    if unmatched_entries:
        return True

    if not loc_sales:
        return False

    totals_to_check = (
        loc_sales.get("total"),
        loc_sales.get("total_amount"),
        loc_sales.get("net_including_tax_total"),
        loc_sales.get("discount_total"),
    )

    return any(value is not None for value in totals_to_check)


event = Blueprint("event", __name__)


def _terminal_sales_serializer() -> URLSafeSerializer:
    secret_key = current_app.secret_key or current_app.config.get("SECRET_KEY")
    if not secret_key:
        raise RuntimeError("Application secret key is not configured.")
    return URLSafeSerializer(secret_key, salt="terminal-sales-resolution")


_TERMINAL_SALES_STATE_KEY = "terminal_sales_state"


def _parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _get_event_filters(source):
    raw_date_filter = (source.get("date_filter") or "").strip()
    date_filter_value = ""
    parsed_date_filter = _parse_date(raw_date_filter)
    if parsed_date_filter:
        date_filter_value = parsed_date_filter.isoformat()
    return {
        "type": (source.get("type") or "").strip(),
        "name_contains": (source.get("name_contains") or "").strip(),
        "name_not_contains": (source.get("name_not_contains") or "").strip(),
        "start_date_from": (source.get("start_date_from") or "").strip(),
        "start_date_to": (source.get("start_date_to") or "").strip(),
        "end_date_from": (source.get("end_date_from") or "").strip(),
        "end_date_to": (source.get("end_date_to") or "").strip(),
        "closed_status": (source.get("closed_status") or "").strip(),
        "date_filter": date_filter_value,
    }


def _apply_event_filters(query, filters):
    event_type = filters.get("type")
    if event_type:
        query = query.filter_by(event_type=event_type)

    name_contains = filters.get("name_contains")
    if name_contains:
        query = query.filter(
            build_text_match_predicate(Event.name, name_contains, "contains")
        )

    name_not_contains = filters.get("name_not_contains")
    if name_not_contains:
        query = query.filter(
            build_text_match_predicate(
                Event.name, name_not_contains, "not_contains"
            )
        )

    start_date_from = _parse_date(filters.get("start_date_from"))
    if start_date_from:
        query = query.filter(Event.start_date >= start_date_from)

    start_date_to = _parse_date(filters.get("start_date_to"))
    if start_date_to:
        query = query.filter(Event.start_date <= start_date_to)

    end_date_from = _parse_date(filters.get("end_date_from"))
    if end_date_from:
        query = query.filter(Event.end_date >= end_date_from)

    end_date_to = _parse_date(filters.get("end_date_to"))
    if end_date_to:
        query = query.filter(Event.end_date <= end_date_to)

    date_filter = _parse_date(filters.get("date_filter"))
    if date_filter:
        query = query.filter(
            Event.start_date <= date_filter, Event.end_date >= date_filter
        )

    closed_status = filters.get("closed_status")
    if closed_status == "open":
        query = query.filter(Event.closed.is_(False))
    elif closed_status == "closed":
        query = query.filter(Event.closed.is_(True))

    return query


@event.route("/events")
@login_required
def view_events():
    scope = request.endpoint or "event.view_events"
    default_filters = get_filter_defaults(current_user, scope)
    active_filters = normalize_filters(
        request.args, exclude=("page", "per_page", "reset")
    )
    if default_filters and not active_filters:
        return redirect(
            url_for("event.view_events", **filters_to_query_args(default_filters))
        )

    filters = _get_event_filters(request.args)
    query = _apply_event_filters(Event.query, filters)
    events = query.all()
    create_form = EventForm()
    return render_template(
        "events/view_events.html",
        events=events,
        event_types=EVENT_TYPES,
        type_labels=dict(EVENT_TYPES),
        create_form=create_form,
        filter_values=filters,
    )


@event.route("/events/create", methods=["GET", "POST"])
@login_required
def create_event():
    form = EventForm()
    if form.validate_on_submit():
        ev = Event(
            name=form.name.data,
            start_date=form.start_date.data,
            end_date=form.end_date.data,
            event_type=form.event_type.data,
            estimated_sales=form.estimated_sales.data,
        )
        db.session.add(ev)
        db.session.commit()
        log_activity(f"Created event {ev.id}")
        flash("Event created")
        return redirect(url_for("event.view_events"))
    return render_template(
        "events/create_event.html",
        form=form,
        terminal_sales_conflict_guidance=_TERMINAL_SALES_CONFLICT_GUIDANCE,
    )


@event.route("/events/filter", methods=["POST"])
@login_required
def filter_events_ajax():
    filters = _get_event_filters(request.form)
    events = _apply_event_filters(Event.query, filters).all()
    return render_template(
        "events/_events_table.html",
        events=events,
        type_labels=dict(EVENT_TYPES),
    )


@event.route("/events/create/ajax", methods=["POST"])
@login_required
def create_event_ajax():
    form = EventForm()
    if form.validate_on_submit():
        ev = Event(
            name=form.name.data,
            start_date=form.start_date.data,
            end_date=form.end_date.data,
            event_type=form.event_type.data,
            estimated_sales=form.estimated_sales.data,
        )
        db.session.add(ev)
        db.session.commit()
        log_activity(f"Created event {ev.id}")
        return render_template(
            "events/_event_row.html", e=ev, type_labels=dict(EVENT_TYPES)
        )
    response = {"errors": form.errors or {"form": ["Invalid data submitted."]}}
    return jsonify(response), 400


@event.route("/events/<int:event_id>/edit", methods=["GET", "POST"])
@login_required
def edit_event(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)
    form = EventForm(obj=ev)
    if form.validate_on_submit():
        conflicts = _find_terminal_sales_event_location_conflicts(
            start_date=form.start_date.data,
            end_date=form.end_date.data,
            location_ids=[event_location.location_id for event_location in ev.locations],
            exclude_event_id=ev.id,
        )
        if conflicts:
            message = _build_terminal_sales_conflict_message(conflicts)
            form.start_date.errors.append(message)
            return render_template(
                "events/edit_event.html",
                form=form,
                event=ev,
                terminal_sales_conflict_guidance=_TERMINAL_SALES_CONFLICT_GUIDANCE,
                terminal_sales_conflicts=_terminal_sales_conflicts_for_event(ev),
            )
        ev.name = form.name.data
        ev.start_date = form.start_date.data
        ev.end_date = form.end_date.data
        ev.event_type = form.event_type.data
        ev.estimated_sales = form.estimated_sales.data
        db.session.commit()
        log_activity(f"Edited event {ev.id}")
        flash("Event updated")
        return redirect(url_for("event.view_events"))
    return render_template(
        "events/edit_event.html",
        form=form,
        event=ev,
        terminal_sales_conflict_guidance=_TERMINAL_SALES_CONFLICT_GUIDANCE,
        terminal_sales_conflicts=_terminal_sales_conflicts_for_event(ev),
    )


@event.route("/events/<int:event_id>/delete", methods=["POST"])
@login_required
def delete_event(event_id):
    form = CSRFOnlyForm()
    if not form.validate_on_submit():
        abort(400)
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)
    event_id = ev.id
    db.session.delete(ev)
    db.session.commit()
    log_activity(f"Deleted event {event_id}")
    flash("Event deleted")
    return redirect(url_for("event.view_events"))


@event.route("/events/<int:event_id>")
@login_required
def view_event(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)
    type_labels = dict(EVENT_TYPES)
    opening_form = UpdateOpeningCountsForm()
    opening_form.location_ids.choices = [
        (el.id, el.location.name)
        for el in ev.locations
        if el.location is not None
    ]
    confirmed_sales = _calculate_confirmed_sales_summary(ev)
    physical_terminal_variance = _calculate_physical_vs_terminal_variance(ev)
    terminal_sales_conflicts = _terminal_sales_conflicts_for_event(ev)
    return render_template(
        "events/view_event.html",
        event=ev,
        event_type_label=type_labels.get(ev.event_type, ev.event_type),
        opening_form=opening_form,
        confirmed_sales=confirmed_sales,
        physical_terminal_variance=physical_terminal_variance,
        terminal_sales_conflict_guidance=_TERMINAL_SALES_CONFLICT_GUIDANCE,
        terminal_sales_conflicts=terminal_sales_conflicts,
        terminal_sales_conflicts_by_location=_group_terminal_sales_conflicts_by_location(
            terminal_sales_conflicts
        ),
        undo_confirm_form_factory=EventLocationUndoConfirmForm,
    )


@event.route("/events/<int:event_id>/close-report")
@login_required
def closed_event_report(event_id):
    event = (
        Event.query.options(
            selectinload(Event.locations).selectinload(EventLocation.location),
            selectinload(Event.locations)
            .selectinload(EventLocation.terminal_sales)
            .selectinload(TerminalSale.product),
            selectinload(Event.locations)
            .selectinload(EventLocation.stand_sheet_items)
            .selectinload(EventStandSheetItem.item),
        )
        .filter(Event.id == event_id)
        .first()
    )

    if event is None or not event.closed:
        abort(404)

    conversions = _conversion_mapping()
    location_reports: list[SimpleNamespace] = []
    total_terminal_quantity = 0.0
    total_terminal_amount = Decimal("0.00")
    total_physical_quantity = 0.0
    total_physical_amount = Decimal("0.00")
    has_priced_physical_total = False
    any_sheet_data = False

    for event_location in sorted(
        event.locations,
        key=lambda el: (el.location.name.lower() if el.location else ""),
    ):
        location_obj, stand_items = _get_stand_items(
            event_location.location_id, event.id
        )
        stand_items.sort(
            key=lambda entry: (
                entry.get("item").name.casefold()
                if entry.get("item") is not None
                else ""
            ),
            reverse=True,
        )
        price_lookup = _build_item_price_lookup(event_location, stand_items)

        location_terminal_quantity = 0.0
        location_terminal_amount = Decimal("0.00")
        for sale in event_location.terminal_sales:
            quantity_value = sale.quantity or 0.0
            quantity = float(quantity_value)
            quantity_decimal = Decimal(str(quantity_value or 0.0))
            product = sale.product
            price_decimal = (
                Decimal(str(getattr(product, "price", 0.0) or 0.0))
                if product
                else Decimal("0.00")
            )
            location_terminal_quantity += quantity
            location_terminal_amount += quantity_decimal * price_decimal

        location_physical_quantity = 0.0
        location_physical_amount = Decimal("0.00")
        location_has_priced_physical = False
        location_has_sheet_data = False

        for entry in stand_items:
            item = entry.get("item")
            price_per_unit = (
                price_lookup.get(item.id) if item is not None else None
            )
            entry["price_per_unit"] = price_per_unit

            sheet = entry.get("sheet")
            if sheet is None:
                entry["physical_units"] = None
                entry["physical_units_display"] = None
                entry["physical_amount"] = None
                continue

            location_has_sheet_data = True
            any_sheet_data = True

            opening = float(sheet.opening_count or 0.0)
            transferred_in = float(sheet.transferred_in or 0.0)
            transferred_out = float(sheet.transferred_out or 0.0)
            adjustments = float(sheet.adjustments or 0.0)
            eaten = float(sheet.eaten or 0.0)
            spoiled = float(sheet.spoiled or 0.0)
            closing = float(sheet.closing_count or 0.0)

            physical_units_base = (
                opening
                + transferred_in
                + adjustments
                - transferred_out
                - eaten
                - spoiled
                - closing
            )

            entry["physical_units"] = physical_units_base
            entry["physical_units_display"] = _convert_value_for_reporting(
                physical_units_base, entry.get("base_unit"), conversions
            )

            location_physical_quantity += physical_units_base
            total_physical_quantity += physical_units_base

            if price_per_unit is not None:
                amount_decimal = Decimal(str(physical_units_base)) * Decimal(
                    str(price_per_unit)
                )
                entry["physical_amount"] = _quantize_currency(amount_decimal)
                location_physical_amount += amount_decimal
                location_has_priced_physical = True
            else:
                entry["physical_amount"] = None

        total_terminal_quantity += location_terminal_quantity
        total_terminal_amount += location_terminal_amount

        location_terminal_amount_display = _quantize_currency(
            location_terminal_amount
        )
        location_physical_amount_display = (
            _quantize_currency(location_physical_amount)
            if location_has_priced_physical
            else None
        )

        if location_has_priced_physical:
            total_physical_amount += location_physical_amount
            has_priced_physical_total = True

        location_reports.append(
            SimpleNamespace(
                event_location=event_location,
                location=location_obj,
                stand_items=stand_items,
                has_sheet_data=location_has_sheet_data,
                notes=event_location.notes,
                terminal=SimpleNamespace(
                    quantity=location_terminal_quantity,
                    amount=location_terminal_amount_display,
                ),
                physical=SimpleNamespace(
                    quantity=location_physical_quantity,
                    amount=location_physical_amount_display,
                ),
            )
        )

    totals = SimpleNamespace(
        terminal_quantity=total_terminal_quantity,
        terminal_amount=_quantize_currency(total_terminal_amount),
        physical_quantity=total_physical_quantity,
        physical_amount=
        _quantize_currency(total_physical_amount)
        if has_priced_physical_total
        else None,
    )

    return render_template(
        "events/close_report.html",
        event=event,
        totals=totals,
        locations=location_reports,
        has_stand_data=any_sheet_data,
    )


@event.route(
    "/events/<int:event_id>/update_opening_counts", methods=["POST"]
)
@login_required
def update_opening_counts(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)

    form = UpdateOpeningCountsForm()
    event_locations = (
        EventLocation.query.filter_by(event_id=event_id)
        .join(Location)
        .order_by(Location.name)
        .all()
    )
    form.location_ids.choices = [
        (el.id, el.location.name) for el in event_locations
    ]

    if not form.validate_on_submit():
        flash("Unable to update opening counts. Please try again.", "warning")
        return redirect(url_for("event.view_event", event_id=event_id))

    if ev.closed:
        flash("This event is closed and opening counts cannot be updated.", "warning")
        return redirect(url_for("event.view_event", event_id=event_id))

    selected_ids = form.location_ids.data or []
    if not selected_ids:
        flash("Select at least one location to update opening counts.", "warning")
        return redirect(url_for("event.view_event", event_id=event_id))

    location_map = {el.id: el for el in event_locations}
    updated_names = []
    skipped_names = []
    for el_id in selected_ids:
        el = location_map.get(el_id)
        if el is None:
            continue
        if el.confirmed:
            skipped_names.append(el.location.name)
            continue
        _sync_event_location_opening_counts(el)
        updated_names.append(el.location.name)

    if not updated_names:
        if skipped_names:
            flash(
                "The selected locations are already confirmed and cannot be updated.",
                "warning",
            )
        else:
            flash("No matching locations were found to update.", "warning")
        return redirect(url_for("event.view_event", event_id=event_id))

    db.session.commit()

    log_activity(
        "Updated opening counts for event %s locations: %s"
        % (event_id, ", ".join(updated_names))
    )

    message = "Opening counts updated for: %s" % ", ".join(updated_names)
    if skipped_names:
        message += ". Skipped confirmed locations: %s" % ", ".join(skipped_names)
    flash(message, "success")

    return redirect(url_for("event.view_event", event_id=event_id))


@event.route("/events/<int:event_id>/add_location", methods=["GET", "POST"])
@login_required
def add_location(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)
    form = EventLocationForm(event_id=event_id)
    if not form.location_id.choices:
        flash("All available locations have already been assigned to this event.")
        return redirect(url_for("event.view_event", event_id=event_id))
    if form.validate_on_submit():
        selected_ids = form.location_id.data
        conflicts = _find_terminal_sales_event_location_conflicts(
            start_date=ev.start_date,
            end_date=ev.end_date,
            location_ids=selected_ids,
            exclude_event_id=event_id,
        )
        if conflicts:
            message = _build_terminal_sales_conflict_message(conflicts)
            form.location_id.errors.append(message)
            return render_template(
                "events/add_location.html",
                form=form,
                event=ev,
                terminal_sales_conflict_guidance=_TERMINAL_SALES_CONFLICT_GUIDANCE,
            )
        event_locations = []
        for location_id in selected_ids:
            event_location = EventLocation(
                event_id=event_id, location_id=location_id
            )
            db.session.add(event_location)
            event_locations.append(event_location)
        if event_locations:
            db.session.flush()
            for event_location in event_locations:
                _sync_event_location_opening_counts(event_location)
        db.session.commit()
        location_names = []
        for location_id in selected_ids:
            location = db.session.get(Location, location_id)
            location_names.append(location.name if location else str(location_id))
        location_list = ", ".join(location_names)
        log_activity(
            f"Assigned locations {location_list} to event {event_id}"
        )
        flash(
            "Locations assigned"
            if len(event_locations) > 1
            else "Location assigned"
        )
        return redirect(url_for("event.view_event", event_id=event_id))
    return render_template(
        "events/add_location.html",
        form=form,
        event=ev,
        terminal_sales_conflict_guidance=_TERMINAL_SALES_CONFLICT_GUIDANCE,
    )


@event.route(
    "/events/<int:event_id>/locations/<int:el_id>/sales/add",
    methods=["GET", "POST"],
)
@login_required
def add_terminal_sale(event_id, el_id):
    el = db.session.get(EventLocation, el_id)
    if el is None or el.event_id != event_id:
        abort(404)
    conflicts = _find_terminal_sales_event_location_conflicts(
        start_date=el.event.start_date,
        end_date=el.event.end_date,
        location_ids=[el.location_id],
        exclude_event_id=event_id,
    )
    if conflicts:
        flash(_build_terminal_sales_conflict_message(conflicts), "warning")
        return redirect(url_for("event.view_event", event_id=event_id))
    if el.event.closed:
        flash("This location is closed and cannot accept new sales.")
        return redirect(url_for("event.view_event", event_id=event_id))

    if el.confirmed:
        flash("This location is closed and cannot accept new sales.")
        return redirect(url_for("event.view_event", event_id=event_id))

    def _collect_event_location_products(event_location: EventLocation):
        location_obj = event_location.location
        products: list[Product] = []
        if location_obj is not None:
            products.extend(get_authoritative_location_products(location_obj))
        for sale in event_location.terminal_sales:
            product = sale.product
            if product is None:
                continue
            if product not in products:
                products.append(product)
        products.sort(key=lambda prod: prod.name.lower())
        return products

    available_products = _collect_event_location_products(el)

    if request.method == "POST":
        updated = False
        for product in available_products:
            qty = request.form.get(f"qty_{product.id}")
            try:
                amount = float(qty) if qty else 0
            except ValueError:
                amount = 0

            sale = TerminalSale.query.filter_by(
                event_location_id=el_id, product_id=product.id
            ).first()

            if amount:
                if sale:
                    if sale.quantity != amount:
                        sale.quantity = amount
                        updated = True
                else:
                    sale = TerminalSale(
                        event_location_id=el_id,
                        product_id=product.id,
                        quantity=amount,
                        sold_at=datetime.utcnow(),
                    )
                    db.session.add(sale)
                    updated = True
            elif sale:
                db.session.delete(sale)
                updated = True

        db.session.commit()
        if updated:
            log_activity(f"Updated terminal sales for event location {el_id}")
        flash("Sales recorded")
        return redirect(url_for("event.view_event", event_id=event_id))

    existing_sales = {s.product_id: s.quantity for s in el.terminal_sales}
    return render_template(
        "events/add_terminal_sales.html",
        event_location=el,
        existing_sales=existing_sales,
        products=available_products,
    )


def _wants_json_response() -> bool:
    """Return True when the current request prefers a JSON response."""

    if request.is_json:
        return True
    accept_mimetypes = request.accept_mimetypes
    return (
        accept_mimetypes["application/json"]
        > accept_mimetypes["text/html"]
    )


def _serialize_scan_totals(event_location: EventLocation):
    """Return the location and summaries of counted items."""

    location, stand_items = _get_stand_items(
        event_location.location_id, event_location.event_id
    )
    totals = []
    seen_item_ids = set()

    for entry in stand_items:
        item = entry["item"]
        sheet = entry.get("sheet")
        counted = float(sheet.closing_count or 0.0) if sheet else 0.0
        totals.append(
            {
                "item_id": item.id,
                "name": item.name,
                "upc": item.upc,
                "expected": float(entry.get("expected") or 0.0),
                "counted": counted,
                "base_unit": item.base_unit,
            }
        )
        seen_item_ids.add(item.id)

    for sheet in event_location.stand_sheet_items:
        if sheet.item_id in seen_item_ids:
            continue
        item = sheet.item
        totals.append(
            {
                "item_id": item.id,
                "name": item.name,
                "upc": item.upc,
                "expected": 0.0,
                "counted": float(sheet.closing_count or 0.0),
                "base_unit": item.base_unit,
            }
        )

    totals.sort(key=lambda record: record["name"].lower())
    return location, totals


@event.route(
    "/events/<int:event_id>/locations/<int:location_id>/scan_counts",
    methods=["GET", "POST"],
)
@login_required
def scan_counts(event_id, location_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)
    if ev.event_type != "inventory":
        abort(404)

    el = EventLocation.query.filter_by(
        event_id=event_id, location_id=location_id
    ).first()
    if el is None:
        abort(404)

    wants_json = _wants_json_response()

    if ev.closed:
        if wants_json:
            return (
                jsonify(
                    success=False, error="This event is closed to updates."
                ),
                403,
            )
        abort(403, description="This event is closed to updates.")

    form = ScanCountForm()
    if form.quantity.data is None:
        form.quantity.data = 1

    if request.method == "GET" and wants_json:
        location, totals = _serialize_scan_totals(el)
        return jsonify(
            success=True,
            location={"id": location.id, "name": location.name},
            totals=totals,
        )

    if request.method == "POST":
        if wants_json:
            payload = request.get_json(silent=True) or {}
            upc = (payload.get("upc") or "").strip()
            raw_quantity = payload.get("quantity", 1)
            try:
                quantity = float(raw_quantity)
            except (TypeError, ValueError):
                quantity = None

            if not upc:
                return (
                    jsonify(
                        success=False, error="A barcode value is required."
                    ),
                    400,
                )
            if quantity is None:
                return (
                    jsonify(
                        success=False,
                        error="Quantity must be a numeric value.",
                    ),
                    400,
                )
        else:
            if not form.validate_on_submit():
                location, totals = _serialize_scan_totals(el)
                return (
                    render_template(
                        "events/scan_count.html",
                        event=ev,
                        location=location,
                        form=form,
                        totals=totals,
                    ),
                    400,
                )
            upc = (form.upc.data or "").strip()
            quantity = float(form.quantity.data or 0)

        item = Item.lookup_by_barcode(upc)
        if item is None:
            if wants_json:
                return (
                    jsonify(
                        success=False,
                        error=f"No item found for barcode {upc}.",
                    ),
                    404,
                )
            flash(f"No item found for barcode {upc}.", "danger")
            location, totals = _serialize_scan_totals(el)
            return (
                render_template(
                    "events/scan_count.html",
                    event=ev,
                    location=location,
                    form=form,
                    totals=totals,
                ),
                404,
            )

        sheet = EventStandSheetItem.query.filter_by(
            event_location_id=el.id, item_id=item.id
        ).first()
        if sheet is None:
            sheet = EventStandSheetItem(
                event_location_id=el.id, item_id=item.id
            )
            db.session.add(sheet)

        sheet.transferred_out = (sheet.transferred_out or 0.0) + quantity
        sheet.closing_count = (sheet.closing_count or 0.0) + quantity
        db.session.commit()
        log_activity(
            f"Recorded scan count for event {event_id}"
            f" location {location_id} item {item.id}"
        )

        location, totals = _serialize_scan_totals(el)

        if wants_json:
            return jsonify(
                success=True,
                item={
                    "id": item.id,
                    "name": item.name,
                    "upc": item.upc,
                    "quantity": quantity,
                    "total": float(sheet.transferred_out or 0.0),
                    "base_unit": item.base_unit,
                },
                totals=totals,
            )

        flash(
            f"Recorded {quantity:g} {item.base_unit} for {item.name}.",
            "success",
        )
        return redirect(
            url_for(
                "event.scan_counts",
                event_id=event_id,
                location_id=location_id,
            )
        )

    location, totals = _serialize_scan_totals(el)
    return render_template(
        "events/scan_count.html",
        event=ev,
        location=location,
        form=form,
        totals=totals,
    )


@event.route("/events/<int:event_id>/terminal-sales", methods=["GET", "POST"])
@event.route("/events/<int:event_id>/sales/upload", methods=["GET", "POST"])
@login_required
def upload_terminal_sales(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)
    conflicts = _terminal_sales_conflicts_for_event(ev)
    if conflicts:
        flash(_build_terminal_sales_conflict_message(conflicts), "warning")
        return redirect(url_for("event.view_event", event_id=event_id))

    form = TerminalSalesUploadForm()
    product_form = ProductWithRecipeForm()
    if product_form.recipe_yield_quantity.data is None:
        product_form.recipe_yield_quantity.data = 1
    if hasattr(product_form, "submit"):
        product_form.submit.label.text = "Create Product"
    serializer = _terminal_sales_serializer()
    state_store = session.get(_TERMINAL_SALES_STATE_KEY)
    if not isinstance(state_store, dict):
        state_store = {}
    event_state_key = str(event_id)

    def _serialize_token(token_id: str) -> str:
        return serializer.dumps({"event_id": event_id, "token_id": token_id})

    def _prepare_state_payload(data: dict) -> dict:
        try:
            return json.loads(json.dumps(data))
        except (TypeError, ValueError):
            return data

    def _load_state_row(token_id: str) -> TerminalSalesResolutionState | None:
        if not token_id or not current_user.is_authenticated:
            return None
        return (
            TerminalSalesResolutionState.query.filter_by(
                event_id=event_id,
                user_id=current_user.id,
                token_id=token_id,
            ).one_or_none()
        )

    def _load_state_payload(token_id: str) -> dict | None:
        row = _load_state_row(token_id)
        if row is None or not isinstance(row.payload, dict):
            return None
        payload = dict(row.payload)
        payload.setdefault("token_id", token_id)
        return payload

    def _store_state_payload(token_id: str, payload: dict) -> None:
        if not current_user.is_authenticated:
            return
        bind = db.session.get_bind()
        if bind is None:
            return
        with Session(bind=bind) as state_session:
            state_row = (
                state_session.query(TerminalSalesResolutionState)
                .filter_by(
                    event_id=event_id,
                    user_id=current_user.id,
                    token_id=token_id,
                )
                .one_or_none()
            )
            if state_row is None:
                state_row = TerminalSalesResolutionState(
                    event_id=event_id,
                    user_id=current_user.id,
                    token_id=token_id,
                )
            state_row.payload = payload
            state_session.add(state_row)
            state_session.query(TerminalSalesResolutionState).filter(
                TerminalSalesResolutionState.event_id == event_id,
                TerminalSalesResolutionState.user_id == current_user.id,
                TerminalSalesResolutionState.token_id != token_id,
            ).delete(synchronize_session=False)
            state_session.commit()
        db.session.expire_all()

    def _clear_state() -> None:
        removed = state_store.pop(event_state_key, None)
        if _TERMINAL_SALES_STATE_KEY in session or removed is not None:
            session[_TERMINAL_SALES_STATE_KEY] = state_store
            session.modified = True
        if not current_user.is_authenticated:
            return
        bind = db.session.get_bind()
        if bind is None:
            return
        with Session(bind=bind) as state_session:
            query = state_session.query(TerminalSalesResolutionState).filter_by(
                event_id=event_id,
                user_id=current_user.id,
            )
            token_to_remove = None
            if isinstance(removed, dict):
                token_to_remove = removed.get("token_id")
            elif isinstance(removed, str):
                token_to_remove = removed
            if token_to_remove:
                query = query.filter_by(token_id=token_to_remove)
            query.delete(synchronize_session=False)
            state_session.commit()
        db.session.expire_all()

    state_entry = state_store.get(event_state_key)
    if isinstance(state_entry, str):
        state_entry = {"token_id": state_entry}
    elif not isinstance(state_entry, dict):
        state_entry = None

    stored_token_id = state_entry.get("token_id") if state_entry else None
    state_token: str | None = None
    state_data: dict | None = None

    def _invalidate_state():
        nonlocal state_entry, stored_token_id, state_token, state_data
        _clear_state()
        state_entry = None
        stored_token_id = None
        state_token = None
        state_data = None

    def _save_state(data: dict) -> tuple[str, dict]:
        nonlocal state_entry, stored_token_id
        token_id = data.get("token_id") or token_urlsafe(16)
        data["token_id"] = token_id
        token = _serialize_token(token_id)
        state_store[event_state_key] = {"token_id": token_id}
        session[_TERMINAL_SALES_STATE_KEY] = state_store
        session.modified = True
        stored_payload = _prepare_state_payload(data)
        _store_state_payload(token_id, stored_payload)
        state_entry = {"token_id": token_id}
        stored_token_id = token_id
        return token, data

    posted_state_token = (
        request.form.get("state_token") if request.method == "POST" else None
    )
    queried_state_token = (
        request.args.get("state_token") if request.method != "POST" else None
    )

    if request.method == "POST":
        if posted_state_token:
            try:
                token_payload = serializer.loads(posted_state_token)
            except BadSignature:
                _invalidate_state()
            else:
                token_id = token_payload.get("token_id") if isinstance(token_payload, dict) else None
                token_event_id = token_payload.get("event_id") if isinstance(token_payload, dict) else None
                if token_event_id is not None and token_event_id != event_id:
                    _invalidate_state()
                elif stored_token_id and token_id != stored_token_id:
                    _invalidate_state()
                else:
                    loaded_state = _load_state_payload(token_id or "")
                    if loaded_state is None:
                        if token_id:
                            _clear_state()
                            state_entry = None
                            stored_token_id = None
                            state_data = {"token_id": token_id}
                            state_token = _serialize_token(token_id)
                        else:
                            _invalidate_state()
                    else:
                        state_data = loaded_state
                        state_token = _serialize_token(token_id)
                        if not stored_token_id and token_id:
                            state_store[event_state_key] = {"token_id": token_id}
                            session[_TERMINAL_SALES_STATE_KEY] = state_store
                            session.modified = True
                            state_entry = {"token_id": token_id}
                            stored_token_id = token_id
        elif stored_token_id:
            loaded_state = _load_state_payload(stored_token_id)
            if loaded_state is None:
                _invalidate_state()
            else:
                state_data = loaded_state
                state_token = _serialize_token(stored_token_id)
    else:
        if queried_state_token:
            try:
                token_payload = serializer.loads(queried_state_token)
            except BadSignature:
                _invalidate_state()
            else:
                token_id = (
                    token_payload.get("token_id") if isinstance(token_payload, dict) else None
                )
                token_event_id = (
                    token_payload.get("event_id") if isinstance(token_payload, dict) else None
                )
                if not token_id or (token_event_id is not None and token_event_id != event_id):
                    _invalidate_state()
                else:
                    loaded_state = _load_state_payload(token_id)
                    if loaded_state is None:
                        _invalidate_state()
                    else:
                        state_token, state_data = _save_state(dict(loaded_state))
        else:
            _invalidate_state()
    open_locations = [
        el
        for el in ev.locations
        if not el.confirmed and not ev.closed
    ]
    mapping_payload = None
    mapping_filename = None
    sales_summary: dict[str, dict] = {}
    sales_location_names: list[str] = []
    default_mapping: dict[int, str] = {}
    unresolved_products: list[dict] = []
    resolution_errors: list[str] = []
    product_resolution_required = False
    product_choices: list[Product] = []
    product_search_options: list[dict[str, str]] = []
    CREATE_SELECTION_VALUE = "__create__"
    SKIP_SELECTION_VALUE = "__skip__"
    price_discrepancies: dict[str, list[dict]] = {}
    menu_mismatches: dict[str, list[dict]] = {}
    warnings_required = False
    warnings_acknowledged = False
    ignored_sales_locations: set[str] = set()
    assigned_sales_locations: set[str] = set()
    unassigned_sales_locations: list[str] = []
    assignment_errors: list[str] = []
    active_stage = "upload"
    countable_products: list[dict] = []
    countable_item_options: list[dict[str, str]] = []
    countable_selection_errors: list[str] = []
    countable_stage_requested = (
        request.form.get("countable-selection-step") == "1"
    )
    purchase_gl_codes: list[GLCode] = []
    product_mapping_preview: list[dict] = []
    wizard_stage = "upload"
    current_issue = None
    issue_index = 0
    remaining_locations = 0
    total_locations = 0
    selected_locations: list[str] = []
    created_product_ids_state: set[int] = set()
    product_creations_state: dict[str, int] = {}

    def _get_purchase_gl_codes() -> list[GLCode]:
        nonlocal purchase_gl_codes
        if not purchase_gl_codes:
            purchase_gl_codes = (
                GLCode.query.filter(
                    or_(GLCode.code.like("5%"), GLCode.code.like("6%"))
                )
                .order_by(GLCode.code)
                .all()
            )
        return purchase_gl_codes

    def _normalize_product_name(value: str) -> str:
        return normalize_pos_alias(value)

    def _normalize_location_name(value: str) -> str:
        return normalize_pos_alias(value)

    def _normalized_sql_expression(column):
        lowered = func.lower(column)
        alphanumeric = func.regexp_replace(lowered, r"[^a-z0-9]+", " ", "g")
        compacted = func.regexp_replace(alphanumeric, r"\s+", " ", "g")
        return func.trim(compacted)

    def _group_rows(row_data):
        return group_terminal_sales_rows(row_data)

    def _derive_price_map(summary: dict[str, dict]) -> dict[str, float | None]:
        """Build a mapping of product names to representative sale prices."""

        price_map: dict[str, float | None] = {}
        sentinel = object()
        for data in summary.values():
            products = data.get("products", {})
            for name, details in products.items():
                existing = price_map.get(name, sentinel)
                if existing is not sentinel and existing is not None:
                    # We already derived a concrete price for this name; keep it.
                    continue
                candidate = None
                prices = details.get("prices") or []
                for price in prices:
                    if price is None:
                        continue
                    try:
                        candidate = float(price)
                    except (TypeError, ValueError):
                        continue
                    else:
                        break
                if candidate is None:
                    amount = details.get("amount")
                    quantity = details.get("quantity")
                    try:
                        if amount is not None and quantity:
                            candidate = float(amount) / float(quantity)
                    except (TypeError, ValueError, ZeroDivisionError):
                        candidate = None
                if existing is sentinel:
                    price_map[name] = candidate
                elif existing is None and candidate is not None:
                    price_map[name] = candidate
        return price_map

    def _prices_match(file_price: float, app_price: float) -> bool:
        try:
            return math.isclose(float(file_price), float(app_price), abs_tol=0.01)
        except (TypeError, ValueError):
            return False

    def _terminal_catalog_sell_price(product: Product | None) -> float | None:
        """Return the terminal/event catalog price used for POS reconciliation.

        Guardrail: terminal/event workflows intentionally use Product.price
        (sell price). Product.invoice_sale_price is for customer invoices and
        must not be used here.
        """

        if product is None:
            return None
        return coerce_float(product.price)

    def _store_location_aliases(pending_totals: list[dict]) -> set[str]:
        if not pending_totals:
            return set()

        normalized_entries: dict[str, tuple[str, int]] = {}
        for entry in pending_totals:
            source_name = entry.get("source_location")
            event_location_id = entry.get("event_location_id")
            if not source_name or event_location_id is None:
                continue
            event_location = db.session.get(EventLocation, event_location_id)
            if event_location is None or event_location.location_id is None:
                continue
            normalized = _normalize_location_name(source_name)
            if not normalized:
                continue
            normalized_entries[normalized] = (
                source_name,
                event_location.location_id,
            )

        if not normalized_entries:
            return set()

        existing_aliases = {
            alias.normalized_name: alias
            for alias in TerminalSaleLocationAlias.query.filter(
                TerminalSaleLocationAlias.normalized_name.in_(
                    list(normalized_entries.keys())
                )
            ).all()
        }

        saved_sources: set[str] = set()
        for normalized, (source_name, location_id) in normalized_entries.items():
            alias = existing_aliases.get(normalized)
            if alias is None:
                alias = TerminalSaleLocationAlias(
                    source_name=source_name,
                    normalized_name=normalized,
                    location_id=location_id,
                )
                db.session.add(alias)
            else:
                alias.source_name = source_name
                alias.location_id = location_id
            saved_sources.add(source_name)

        return saved_sources

    if request.method == "POST":
        step = request.form.get("step")
        if step == "resolve":
            active_stage = "menus"
            payload = request.form.get("payload")
            mapping_filename = request.form.get("mapping_filename")
            if not state_token or state_data is None:
                flash("Unable to continue the resolution process.", "danger")
                return redirect(url_for("event.upload_terminal_sales", event_id=event_id))

            token_id = state_data.get("token_id")
            expected_id = (state_entry or {}).get("token_id")
            if not token_id or expected_id != token_id:
                _clear_state()
                flash(
                    "The terminal sales resolution session is no longer valid. "
                    "Upload the sales file again to start over.",
                    "danger",
                )
                return redirect(url_for("event.upload_terminal_sales", event_id=event_id))

            stored_mapping = state_data.get("selected_mapping") or {}
            ignored_sales_locations = set(
                state_data.get("ignored_sales_locations") or []
            )
            assigned_sales_locations = {
                value for value in stored_mapping.values() if value
            }

            queue: list[dict] = state_data.get("queue") or []
            pending_sales: list[dict] = state_data.get("pending_sales") or []
            pending_totals: list[dict] = state_data.get("pending_totals") or []
            selected_locations: list[str] = state_data.get("selected_locations") or []
            menu_candidates: list[dict] = state_data.get("menu_candidates") or []
            menu_candidate_selection = (
                state_data.get("menu_candidate_selection")
                if isinstance(state_data.get("menu_candidate_selection"), dict)
                else {}
            )
            issue_index = state_data.get("issue_index", 0)
            action = request.form.get("action", "")

            if action == "back_to_mapping":
                payload_data = state_data.get("payload") if state_data else None
                if payload_data is None and payload:
                    try:
                        payload_data = json.loads(payload)
                    except (TypeError, ValueError):
                        payload_data = None
                if not payload_data:
                    flash("Unable to process the uploaded sales data.", "danger")
                    return redirect(url_for("event.upload_terminal_sales", event_id=event_id))

                rows = payload_data.get("rows", [])
                mapping_filename = (
                    payload_data.get("filename") or mapping_filename
                )
                if not rows:
                    flash(
                        "No sales records were found in the uploaded file.",
                        "warning",
                    )
                    return redirect(url_for("event.upload_terminal_sales", event_id=event_id))

                sales_summary = _group_rows(rows)
                sales_location_names = list(sales_summary.keys())
                default_mapping: dict[int, str] = {}
                for key, value in stored_mapping.items():
                    try:
                        default_mapping[int(key)] = value
                    except (TypeError, ValueError):
                        continue
                assigned_locations = {
                    value for value in default_mapping.values() if value
                }
                unassigned_sales_locations = sorted(
                    [
                        name
                        for name in sales_location_names
                        if name not in assigned_locations
                        and name not in ignored_sales_locations
                    ]
                )

                state_data["stage"] = "locations"
                state_data["queue"] = []
                state_data["pending_sales"] = []
                state_data["pending_totals"] = []
                state_data["selected_locations"] = []
                state_data["issue_index"] = 0
                state_data["ignored_sales_locations"] = sorted(ignored_sales_locations)
                state_data["selected_mapping"] = stored_mapping
                state_token, state_data = _save_state(state_data)

                return render_template(
                    "events/upload_terminal_sales.html",
                    form=form,
                    event=ev,
                    open_locations=open_locations,
                    mapping_payload=json.dumps(payload_data),
                    mapping_filename=mapping_filename,
                    sales_summary=sales_summary,
                    sales_location_names=sales_location_names,
                    default_mapping=default_mapping,
                    unresolved_products=[],
                    product_choices=product_choices,
                    product_search_options=product_search_options,
                    skip_selection_value=SKIP_SELECTION_VALUE,
                    create_selection_value=CREATE_SELECTION_VALUE,
                    resolution_errors=[],
                    product_resolution_required=False,
                    price_discrepancies={},
                    menu_mismatches={},
                    warnings_required=False,
                    warnings_acknowledged=False,
                    state_token=state_token,
                    ignored_sales_locations=sorted(ignored_sales_locations),
                    assigned_sales_locations=sorted(assigned_locations),
                    unassigned_sales_locations=unassigned_sales_locations,
                    assignment_errors=assignment_errors,
                    product_mapping_preview=product_mapping_preview,
                    active_stage="locations",
                    product_form=product_form,
                    created_product_ids=sorted(created_product_ids_state),
                    wizard_stage="locations",
                )

            if issue_index < 0:
                issue_index = 0
            if issue_index > len(queue):
                issue_index = len(queue)

            if queue and issue_index < len(queue):
                current_issue = queue[issue_index]
            else:
                current_issue = None

            error_messages: list[str] = []

            if action.startswith("price:") and current_issue:
                parts = action.split(":", 2)
                if len(parts) == 3:
                    _, product_id_raw, resolution_value = parts
                    try:
                        product_id_int = int(product_id_raw)
                    except (TypeError, ValueError):
                        error_messages.append("Invalid price resolution request.")
                    else:
                        for issue in current_issue.get("price_issues", []):
                            if issue.get("product_id") == product_id_int:
                                product = db.session.get(Product, product_id_int)
                                if resolution_value == "update":
                                    new_price = issue.get("terminal_price")
                                    if new_price is None:
                                        new_price = issue.get("target_price")
                                    new_price = coerce_float(new_price)
                                    if new_price is None:
                                        error_messages.append(
                                            "Terminal price information is not available for this product."
                                        )
                                        break
                                    if product is None:
                                        error_messages.append(
                                            "Unable to load the product for the selected price resolution."
                                        )
                                        break
                                    issue["resolution"] = "update"
                                    issue["selected_option"] = "terminal"
                                    issue["selected_price"] = new_price
                                    issue["target_price"] = new_price
                                    if issue.get("catalog_price") is None:
                                        issue["catalog_price"] = coerce_float(
                                            _terminal_catalog_sell_price(product)
                                        )
                                    product.price = new_price
                                elif resolution_value == "skip":
                                    if product is None:
                                        error_messages.append(
                                            "Unable to load the product for the selected price resolution."
                                        )
                                        break
                                    catalog_price = issue.get("catalog_price")
                                    catalog_price = coerce_float(catalog_price)
                                    issue["resolution"] = "skip"
                                    issue["selected_option"] = "catalog"
                                    issue["selected_price"] = catalog_price
                                    if catalog_price is None:
                                        product.price = None
                                    else:
                                        product.price = catalog_price
                                break
                else:
                    error_messages.append("Invalid price resolution request.")
            elif action.startswith("menu:") and current_issue:
                parts = action.split(":", 2)
                if len(parts) == 3:
                    _, product_id_raw, resolution_value = parts
                    try:
                        product_id_int = int(product_id_raw)
                    except (TypeError, ValueError):
                        error_messages.append("Invalid menu resolution request.")
                    else:
                        for issue in current_issue.get("menu_issues", []):
                            if issue.get("product_id") == product_id_int:
                                if resolution_value == "add":
                                    issue["resolution"] = "add"
                                elif resolution_value == "skip":
                                    issue["resolution"] = "skip"
                                break
                else:
                    error_messages.append("Invalid menu resolution request.")
            elif action == "next_location":
                if current_issue:
                    unresolved = [
                        issue
                        for issue in current_issue.get("price_issues", [])
                        if issue.get("resolution") is None
                    ]
                    unresolved.extend(
                        issue
                        for issue in current_issue.get("menu_issues", [])
                        if issue.get("resolution") is None
                    )
                    if unresolved:
                        error_messages.append(
                            "Resolve all issues for this location before continuing."
                        )
                    else:
                        issue_index += 1
            elif action == "finish":
                unresolved_overall = []
                for location_issue in queue:
                    unresolved_overall.extend(
                        issue
                        for issue in location_issue.get("price_issues", [])
                        if issue.get("resolution") is None
                    )
                    unresolved_overall.extend(
                        issue
                        for issue in location_issue.get("menu_issues", [])
                        if issue.get("resolution") is None
                    )
                if unresolved_overall:
                    error_messages.append(
                        "Resolve all issues before finishing the import."
                    )
                else:
                    issue_index = len(queue)

            if queue and issue_index < len(queue):
                current_issue = queue[issue_index]
            else:
                current_issue = None

            if issue_index >= len(queue):
                if menu_candidates:
                    state_data["queue"] = queue
                    state_data["pending_sales"] = pending_sales
                    state_data["pending_totals"] = pending_totals
                    state_data["selected_locations"] = selected_locations
                    state_data["issue_index"] = issue_index
                    state_data["stage"] = "menus"
                    state_token, state_data = _save_state(state_data)
                    total_locations = len(queue)
                    active_stage = "menus"
                    return render_template(
                        "events/upload_terminal_sales.html",
                        form=form,
                        event=ev,
                        open_locations=open_locations,
                        mapping_payload=payload,
                        mapping_filename=mapping_filename,
                        sales_summary={},
                        sales_location_names=[],
                        default_mapping={},
                        unresolved_products=[],
                        product_choices=[],
                        product_search_options=product_search_options,
                        skip_selection_value=SKIP_SELECTION_VALUE,
                        create_selection_value=CREATE_SELECTION_VALUE,
                        resolution_errors=resolution_errors,
                        product_resolution_required=False,
                        price_discrepancies={},
                        menu_mismatches={},
                        warnings_required=False,
                        warnings_acknowledged=False,
                        state_token=state_token,
                        issue_index=issue_index,
                        current_issue=None,
                        remaining_locations=0,
                        selected_locations=selected_locations,
                        issue_total=total_locations,
                        menu_candidates=menu_candidates,
                        menu_candidate_selection=menu_candidate_selection,
                        ignored_sales_locations=sorted(ignored_sales_locations),
                        assigned_sales_locations=sorted(assigned_sales_locations),
                        unassigned_sales_locations=unassigned_sales_locations,
                        assignment_errors=assignment_errors,
                        product_mapping_preview=product_mapping_preview,
                        active_stage=active_stage,
                        product_form=product_form,
                        created_product_ids=sorted(created_product_ids_state),
                        wizard_stage="menus",
                    )
                updated_locations = _apply_pending_sales(
                    pending_sales,
                    pending_totals,
                    link_products_to_locations=True,
                )
                saved_location_aliases = _store_location_aliases(pending_totals)
                price_updates, menu_updates = _apply_resolution_actions(
                    {"queue": queue}
                )
                if (
                    updated_locations
                    or price_updates
                    or menu_updates
                    or saved_location_aliases
                ):
                    db.session.commit()
                    log_activity(
                        "Uploaded terminal sales for event "
                        f"{event_id} from {mapping_filename or 'uploaded file'}"
                    )
                    success_parts: list[str] = []
                    if updated_locations:
                        success_parts.append(
                            "Terminal sales were imported for: "
                            + ", ".join(sorted(updated_locations))
                        )
                    if price_updates:
                        success_parts.append(
                            "Updated product prices: " + ", ".join(sorted(set(price_updates)))
                        )
                    if menu_updates:
                        success_parts.append(
                            "Added products to menus: " + ", ".join(sorted(set(menu_updates)))
                        )
                    if saved_location_aliases:
                        success_parts.append(
                            "Remembered location mappings for: "
                            + ", ".join(sorted(saved_location_aliases))
                        )
                    flash(" ".join(success_parts), "success")
                else:
                    flash(
                        "No event locations were linked to the uploaded sales data.",
                        "warning",
                    )
                _clear_state()
                return redirect(url_for("event.view_event", event_id=event_id))

            if error_messages:
                for message in error_messages:
                    flash(message, "danger")

            state_data["queue"] = queue
            state_data["pending_sales"] = pending_sales
            state_data["pending_totals"] = pending_totals
            state_data["selected_locations"] = selected_locations
            state_data["issue_index"] = issue_index
            state_data["token_id"] = token_id
            state_data["ignored_sales_locations"] = sorted(ignored_sales_locations)
            state_data["selected_mapping"] = stored_mapping
            state_data["stage"] = "menus"
            state_token, state_data = _save_state(state_data)

            total_locations = len(queue)
            return render_template(
                    "events/upload_terminal_sales.html",
                    form=form,
                    event=ev,
                    open_locations=open_locations,
                    mapping_payload=payload,
                    mapping_filename=mapping_filename,
                    sales_summary={},
                    sales_location_names=[],
                    default_mapping={},
                    unresolved_products=[],
                    product_choices=[],
                    product_search_options=product_search_options,
                    skip_selection_value=SKIP_SELECTION_VALUE,
                    create_selection_value=CREATE_SELECTION_VALUE,
                    resolution_errors=[],
                    product_resolution_required=False,
                    price_discrepancies={},
                    menu_mismatches={},
                    warnings_required=False,
                    warnings_acknowledged=False,
                    state_token=state_token,
                    issue_index=issue_index,
                    current_issue=current_issue,
                    remaining_locations=len(queue) - issue_index - 1,
                    selected_locations=selected_locations,
                    issue_total=total_locations,
                    ignored_sales_locations=sorted(ignored_sales_locations),
                    assigned_sales_locations=sorted(assigned_sales_locations),
                    unassigned_sales_locations=unassigned_sales_locations,
                    assignment_errors=assignment_errors,
                    product_mapping_preview=product_mapping_preview,
                    active_stage=active_stage,
                    product_form=product_form,
                    created_product_ids=sorted(created_product_ids_state),
                    wizard_stage="menus",
                )

        elif step == "confirm_menus":
            if not state_token or state_data is None:
                flash("Unable to continue the resolution process.", "danger")
                return redirect(url_for("event.upload_terminal_sales", event_id=event_id))

            action = request.form.get("action") or "finish"
            menu_candidates = state_data.get("menu_candidates") or []
            existing_selection = (
                state_data.get("menu_candidate_selection")
                if isinstance(state_data.get("menu_candidate_selection"), dict)
                else {}
            )
            selected_keys = set(request.form.getlist("menu_additions"))
            updated_selection: dict[str, bool] = {}
            for candidate in menu_candidates:
                location_id = candidate.get("event_location_id")
                for product in candidate.get("products", []):
                    key = f"{location_id}:{product.get('product_id')}"
                    updated_selection[key] = key in selected_keys

            if isinstance(state_data, dict):
                state_data["menu_candidate_selection"] = updated_selection
                state_token, state_data = _save_state(state_data)

            if action != "finish":
                return render_template(
                    "events/upload_terminal_sales.html",
                    form=form,
                    event=ev,
                    open_locations=open_locations,
                    mapping_payload=None,
                    mapping_filename=state_data.get("mapping_filename"),
                    sales_summary={},
                    sales_location_names=[],
                    default_mapping={},
                    unresolved_products=[],
                    product_choices=[],
                    product_search_options=product_search_options,
                    skip_selection_value=SKIP_SELECTION_VALUE,
                    create_selection_value=CREATE_SELECTION_VALUE,
                    resolution_errors=[],
                    product_resolution_required=False,
                    price_discrepancies={},
                    menu_mismatches={},
                    warnings_required=False,
                    warnings_acknowledged=False,
                    state_token=state_token,
                    issue_index=len(state_data.get("queue") or []),
                    current_issue=None,
                    remaining_locations=0,
                    selected_locations=state_data.get("selected_locations") or [],
                    issue_total=len(state_data.get("queue") or []),
                    menu_candidates=menu_candidates,
                    menu_candidate_selection=updated_selection,
                    ignored_sales_locations=state_data.get("ignored_sales_locations")
                    or [],
                    assigned_sales_locations=[],
                    unassigned_sales_locations=[],
                    assignment_errors=[],
                    product_mapping_preview=[],
                    active_stage="menus",
                    product_form=product_form,
                    created_product_ids=sorted(created_product_ids_state),
                    wizard_stage="menus",
                )

            pending_sales = state_data.get("pending_sales") or []
            pending_totals = state_data.get("pending_totals") or []
            queue = state_data.get("queue") or []

            combined_queue = list(queue)
            if menu_candidates:
                for candidate in menu_candidates:
                    location_id = candidate.get("event_location_id")
                    menu_issues: list[dict] = []
                    sales_location_name = None
                    for product in candidate.get("products", []):
                        key = f"{location_id}:{product.get('product_id')}"
                        should_add = updated_selection.get(key, False)
                        resolution_value = "add" if should_add else "skip"
                        menu_issues.append(
                            {
                                "product_id": product.get("product_id"),
                                "product": product.get("product_name"),
                                "menu_name": candidate.get("menu_name"),
                                "resolution": resolution_value,
                            }
                        )
                        if sales_location_name is None:
                            sales_location_name = product.get("sales_location")
                    combined_queue.append(
                        {
                            "event_location_id": location_id,
                            "location_name": candidate.get("location_name"),
                            "sales_location": sales_location_name,
                            "price_issues": [],
                            "menu_issues": menu_issues,
                        }
                    )

            updated_locations = _apply_pending_sales(
                pending_sales,
                pending_totals,
                link_products_to_locations=True,
            )
            saved_location_aliases = _store_location_aliases(pending_totals)
            price_updates, menu_updates = _apply_resolution_actions(
                {"queue": combined_queue}
            )
            if (
                updated_locations
                or price_updates
                or menu_updates
                or saved_location_aliases
            ):
                db.session.commit()
                log_activity(
                    "Uploaded terminal sales for event "
                    f"{event_id} from {state_data.get('mapping_filename') or 'uploaded file'}"
                )
                success_parts: list[str] = []
                if updated_locations:
                    success_parts.append(
                        "Terminal sales were imported for: "
                        + ", ".join(sorted(updated_locations))
                    )
                if price_updates:
                    success_parts.append(
                        "Updated product prices: "
                        + ", ".join(sorted(set(price_updates)))
                    )
                if menu_updates:
                    success_parts.append(
                        "Added products to menus: "
                        + ", ".join(sorted(set(menu_updates)))
                    )
                if saved_location_aliases:
                    success_parts.append(
                        "Remembered location mappings for: "
                        + ", ".join(sorted(saved_location_aliases))
                    )
                flash(" ".join(success_parts), "success")
            else:
                flash(
                    "No event locations were linked to the uploaded sales data.",
                    "warning",
                )
            _clear_state()
            return redirect(url_for("event.view_event", event_id=event_id))

        elif step == "map":
            payload = request.form.get("payload")
            payload_data = None
            if payload:
                try:
                    payload_data = json.loads(payload)
                except (TypeError, ValueError):
                    payload_data = None
            if payload_data is None and state_data:
                payload_data = state_data.get("payload")
                if payload_data:
                    payload = json.dumps(payload_data)
            if not payload_data:
                flash("Unable to process the uploaded sales data.", "danger")
                return redirect(url_for("event.upload_terminal_sales", event_id=event_id))

            rows = payload_data.get("rows", [])
            mapping_filename = payload_data.get("filename")
            if not rows:
                flash("No sales records were found in the uploaded file.", "warning")
                return redirect(url_for("event.upload_terminal_sales", event_id=event_id))

            sales_summary = _group_rows(rows)

            stage = request.form.get("stage")
            if not stage and request.form.get("product-resolution-step"):
                stage = "products"
            stage = stage or "locations"
            navigate = request.form.get("navigate") or ""
            active_stage = stage if stage in {"locations", "products"} else "locations"

            ignored_sales_locations = set(request.form.getlist("ignored_locations"))

            selected_mapping = {
                el.id: request.form.get(f"mapping-{el.id}", "")
                for el in open_locations
            }
            assigned_sales_locations = {
                value for value in selected_mapping.values() if value
            }

            state_data = dict(state_data or {})
            stored_mapping = {
                str(el.id): selected_mapping.get(el.id, "")
                for el in open_locations
            }
            state_data["payload"] = payload_data
            state_data["mapping_filename"] = mapping_filename
            state_data["selected_mapping"] = stored_mapping
            state_data["ignored_sales_locations"] = sorted(ignored_sales_locations)
            created_product_ids = set(created_product_ids_state)
            created_product_ids_form: set[int] = set()

            if request.method == "POST":
                for raw_created in request.form.getlist("created_product_ids"):
                    try:
                        created_id = int(raw_created)
                    except (TypeError, ValueError):
                        continue
                    created_product_ids_form.add(created_id)
                    created_product_ids.add(created_id)

            if state_data is not None:
                state_data["created_product_ids"] = sorted(created_product_ids)

            created_product_ids_state = set(created_product_ids)

            conflicting_selections = sorted(
                assigned_sales_locations.intersection(ignored_sales_locations)
            )
            if conflicting_selections:
                assignment_errors.append(
                    "Remove the ignore selection for locations that are also linked: "
                    + ", ".join(conflicting_selections)
                )

            active_sales_summary = {
                name: data
                for name, data in sales_summary.items()
                if name not in ignored_sales_locations
            }

            unassigned_sales_locations = [
                name
                for name in sales_summary.keys()
                if name not in assigned_sales_locations
                and name not in ignored_sales_locations
            ]
            product_price_lookup = _derive_price_map(active_sales_summary)

            product_names = {
                product_name
                for data in active_sales_summary.values()
                for product_name in data["products"].keys()
            }
            product_lookup: dict[str, Product] = {}
            normalized_lookup = {
                name: _normalize_product_name(name) for name in product_names
            }
            created_product_ids = set(created_product_ids_state)

            if product_names:
                product_lookup.update(
                    {
                        p.name: p
                        for p in Product.query.filter(
                            Product.name.in_(product_names)
                        ).all()
                    }
                )

                normalized_values = [
                    norm for norm in normalized_lookup.values() if norm
                ]
                alias_lookup: dict[str, TerminalSaleProductAlias] = {}
                normalized_product_candidates: dict[str, set[Product]] = {}
                if normalized_values:
                    alias_rows = (
                        TerminalSaleProductAlias.query.filter(
                            TerminalSaleProductAlias.normalized_name.in_(
                                normalized_values
                            )
                        ).all()
                    )
                    alias_lookup = {
                        alias.normalized_name: alias for alias in alias_rows
                    }
                    for original_name, normalized in normalized_lookup.items():
                        if (
                            normalized
                            and original_name not in product_lookup
                            and normalized in alias_lookup
                        ):
                            product = alias_lookup[normalized].product
                            if product is not None:
                                product_lookup[original_name] = product

                    normalized_expression = _normalized_sql_expression(
                        Product.name
                    ).label("normalized_name")
                    for product, normalized_name in (
                        Product.query.add_columns(normalized_expression)
                        .filter(normalized_expression.in_(normalized_values))
                        .all()
                    ):
                        if not normalized_name:
                            continue
                        normalized_product_candidates.setdefault(
                            normalized_name, set()
                        ).add(product)
                    if normalized_product_candidates:
                        for original_name, normalized in normalized_lookup.items():
                            if not normalized or original_name in product_lookup:
                                continue
                            candidates = normalized_product_candidates.get(
                                normalized
                            )
                            if not candidates or len(candidates) != 1:
                                continue
                            product_lookup[original_name] = next(
                                iter(candidates)
                            )
            else:
                alias_lookup = {}

            unmatched_names = [
                name
                for name in product_names
                if product_lookup.get(name) is None
            ]

            if assignment_errors:
                active_stage = "locations"
                state_data["stage"] = "locations"
                state_token, state_data = _save_state(state_data)
                return render_template(
                    "events/upload_terminal_sales.html",
                    form=form,
                    event=ev,
                    open_locations=open_locations,
                    mapping_payload=payload,
                    mapping_filename=mapping_filename,
                    sales_summary=sales_summary,
                    sales_location_names=list(sales_summary.keys()),
                    default_mapping=selected_mapping,
                    unresolved_products=[],
                    product_choices=[],
                    product_search_options=product_search_options,
                    skip_selection_value=SKIP_SELECTION_VALUE,
                    create_selection_value=CREATE_SELECTION_VALUE,
                    resolution_errors=resolution_errors,
                    product_resolution_required=False,
                    price_discrepancies=price_discrepancies,
                    menu_mismatches=menu_mismatches,
                    warnings_required=warnings_required,
                    warnings_acknowledged=warnings_acknowledged,
                    state_token=state_token,
                    ignored_sales_locations=sorted(ignored_sales_locations),
                    assigned_sales_locations=sorted(assigned_sales_locations),
                    unassigned_sales_locations=unassigned_sales_locations,
                    assignment_errors=assignment_errors,
                    product_mapping_preview=product_mapping_preview,
                    active_stage=active_stage,
                    product_form=product_form,
                    created_product_ids=sorted(created_product_ids_state),
                    wizard_stage="locations",
                )

            if unmatched_names:
                product_resolution_required = True
                resolution_requested = (
                    stage == "products"
                    and navigate != "back"
                    and request.form.get("product-resolution-step") == "1"
                )
                if not product_choices:
                    product_choices = Product.query.order_by(Product.name).all()
                product_search_options = [
                    {
                        "id": str(product.id),
                        "value": f"{product.name} (ID: {product.id})",
                        "label": product.name,
                    }
                    for product in product_choices
                ]

                manual_mappings: dict[str, Product] = {}
                skipped_products: list[str] = []
                product_selections_state: dict[str, str] = {}
                created_product_map = dict(
                    (state_data.get("product_creations") or {})
                )
                if not created_product_map:
                    created_product_map = {}

                for idx, original_name in enumerate(unmatched_names):
                    field_name = f"product-match-{idx}"
                    selected_value = request.form.get(field_name)
                    selected_product = None
                    skip_selected = selected_value == SKIP_SELECTION_VALUE
                    product_selections_state[original_name] = selected_value or ""

                    if selected_value:
                        if skip_selected:
                            skipped_products.append(original_name)
                            created_product_map.pop(original_name, None)
                        elif selected_value == CREATE_SELECTION_VALUE:
                            resolution_errors.append(
                                f"Create the product for '{original_name}' before continuing."
                            )
                            created_product_map.pop(original_name, None)
                        else:
                            try:
                                product_id = int(selected_value)
                            except (TypeError, ValueError):
                                resolution_errors.append(
                                    f"Invalid product selection for {original_name}."
                                )
                                created_product_map.pop(original_name, None)
                            else:
                                selected_product = db.session.get(
                                    Product, product_id
                                )
                                if selected_product is None:
                                    resolution_errors.append(
                                        f"Selected product is no longer available for {original_name}."
                                    )
                                    created_product_map.pop(original_name, None)
                                else:
                                    product_lookup[original_name] = selected_product
                                    manual_mappings[original_name] = selected_product
                                    if (
                                        product_id in created_product_ids_form
                                        or product_id
                                        in (state_data.get("created_product_ids") or [])
                                    ):
                                        created_product_map[original_name] = product_id
                                    else:
                                        created_product_map.pop(original_name, None)
                    elif resolution_requested:
                        resolution_errors.append(
                            f"Select a product or skip '{original_name}' to continue."
                        )
                        created_product_map.pop(original_name, None)
                    else:
                        created_product_map.pop(original_name, None)

                    unresolved_products.append(
                        {
                            "field": field_name,
                            "name": original_name,
                            "selected": selected_value or "",
                            "price": product_price_lookup.get(original_name),
                            "created_product_id": created_product_map.get(
                                original_name
                            ),
                        }
                    )

                state_data["product_selections"] = product_selections_state
                pending_creations = [
                    name
                    for name, product_id in created_product_map.items()
                    if product_id
                ]
                created_product_ids = {
                    product_id
                    for product_id in created_product_map.values()
                    if product_id
                }
                state_data["created_product_ids"] = sorted(created_product_ids)
                state_data["product_creations"] = {
                    name: product_id
                    for name, product_id in created_product_map.items()
                    if product_id
                }
                created_product_ids_state = set(created_product_ids)
                product_creations_state = {
                    name: product_id
                    for name, product_id in created_product_map.items()
                    if product_id
                }

                if not resolution_requested:
                    active_stage = (
                        "locations" if navigate == "back" else "products"
                    )
                    state_data["stage"] = active_stage
                    state_token, state_data = _save_state(state_data)
                    return render_template(
                        "events/upload_terminal_sales.html",
                        form=form,
                        event=ev,
                        open_locations=open_locations,
                        mapping_payload=payload,
                        mapping_filename=mapping_filename,
                        sales_summary=sales_summary,
                        sales_location_names=list(sales_summary.keys()),
                        default_mapping=selected_mapping,
                        unresolved_products=unresolved_products,
                        product_choices=product_choices,
                        product_search_options=product_search_options,
                        skip_selection_value=SKIP_SELECTION_VALUE,
                        create_selection_value=CREATE_SELECTION_VALUE,
                        resolution_errors=resolution_errors,
                        product_resolution_required=True,
                        price_discrepancies=price_discrepancies,
                        menu_mismatches=menu_mismatches,
                        warnings_required=warnings_required,
                        warnings_acknowledged=warnings_acknowledged,
                        state_token=state_token,
                        ignored_sales_locations=sorted(ignored_sales_locations),
                        assigned_sales_locations=sorted(assigned_sales_locations),
                        unassigned_sales_locations=unassigned_sales_locations,
                        assignment_errors=assignment_errors,
                        product_mapping_preview=product_mapping_preview,
                        active_stage=active_stage,
                        product_form=product_form,
                        created_product_ids=sorted(created_product_ids),
                        wizard_stage="products",
                    )

                if (
                    len(manual_mappings) + len(skipped_products)
                    != len(unmatched_names)
                ):
                    resolution_errors.append(
                        "Select a product or skip each unmatched entry to continue."
                    )

                if resolution_errors:
                    active_stage = "products"
                    state_data["stage"] = "products"
                    state_token, state_data = _save_state(state_data)
                    return render_template(
                        "events/upload_terminal_sales.html",
                        form=form,
                        event=ev,
                        open_locations=open_locations,
                        mapping_payload=payload,
                        mapping_filename=mapping_filename,
                        sales_summary=sales_summary,
                        sales_location_names=list(sales_summary.keys()),
                        default_mapping=selected_mapping,
                        unresolved_products=unresolved_products,
                        product_choices=product_choices,
                        product_search_options=product_search_options,
                        skip_selection_value=SKIP_SELECTION_VALUE,
                        create_selection_value=CREATE_SELECTION_VALUE,
                        resolution_errors=resolution_errors,
                        product_resolution_required=True,
                        price_discrepancies=price_discrepancies,
                        menu_mismatches=menu_mismatches,
                        warnings_required=warnings_required,
                        warnings_acknowledged=warnings_acknowledged,
                        state_token=state_token,
                        ignored_sales_locations=sorted(ignored_sales_locations),
                        assigned_sales_locations=sorted(assigned_sales_locations),
                        unassigned_sales_locations=unassigned_sales_locations,
                        assignment_errors=assignment_errors,
                        product_mapping_preview=product_mapping_preview,
                        active_stage=active_stage,
                        product_form=product_form,
                        created_product_ids=sorted(created_product_ids),
                        wizard_stage="products",
                    )

                if pending_creations:
                    pending_recipe_links: list[tuple[Product, Item]] = []
                    countable_targets: list[tuple[str, Product, str, str]] = []
                    countable_interaction = countable_stage_requested

                    for original_name in pending_creations:
                        product = manual_mappings.get(original_name)
                        if product is None:
                            continue
                        field_name = f"countable-item-{product.id}"
                        selected_raw = (request.form.get(field_name) or "").strip()
                        if selected_raw:
                            countable_interaction = True
                        countable_targets.append(
                            (
                                original_name,
                                product,
                                field_name,
                                selected_raw,
                            )
                        )

                    if countable_interaction and countable_targets:
                        item_options = (
                            Item.query.filter_by(archived=False)
                            .order_by(Item.name)
                            .all()
                        )
                        item_lookup = {str(item.id): item for item in item_options}
                        countable_item_options = [
                            {
                                "id": str(item.id),
                                "label": f"{item.name} ({get_unit_label(item.base_unit)})",
                            }
                            for item in item_options
                        ]

                        for original_name, product, field_name, selected_raw in countable_targets:
                            selected_item = (
                                item_lookup.get(selected_raw) if selected_raw else None
                            )
                            countable_products.append(
                                {
                                    "product_id": product.id,
                                    "product_name": product.name,
                                    "field": field_name,
                                    "selected": selected_raw,
                                }
                            )
                            if selected_item is None:
                                if selected_raw and selected_raw not in item_lookup:
                                    countable_selection_errors.append(
                                        "The selected countable item is no longer available "
                                        f"for {product.name}."
                                    )
                                elif countable_stage_requested:
                                    countable_selection_errors.append(
                                        "Select a countable item to track inventory for "
                                        f"{product.name}."
                                    )
                            else:
                                pending_recipe_links.append((product, selected_item))

                        expected_links = len(pending_recipe_links)
                        required_links = len(countable_targets)

                        if (
                            countable_selection_errors
                            or expected_links != required_links
                        ):
                            active_stage = "products"
                            state_data["stage"] = "products"
                            state_token, state_data = _save_state(state_data)
                            if (
                                countable_stage_requested
                                and not countable_selection_errors
                            ):
                                countable_selection_errors.append(
                                    "Choose a countable item for each newly created product before continuing."
                                )
                            return render_template(
                                "events/upload_terminal_sales.html",
                                form=form,
                                event=ev,
                                open_locations=open_locations,
                                mapping_payload=payload,
                                mapping_filename=mapping_filename,
                                sales_summary=sales_summary,
                                sales_location_names=list(sales_summary.keys()),
                                default_mapping=selected_mapping,
                                unresolved_products=unresolved_products,
                                product_choices=product_choices,
                                product_search_options=product_search_options,
                                skip_selection_value=SKIP_SELECTION_VALUE,
                                create_selection_value=CREATE_SELECTION_VALUE,
                                resolution_errors=resolution_errors,
                                product_resolution_required=True,
                                price_discrepancies=price_discrepancies,
                                menu_mismatches=menu_mismatches,
                                warnings_required=warnings_required,
                                warnings_acknowledged=warnings_acknowledged,
                                state_token=state_token,
                                ignored_sales_locations=sorted(ignored_sales_locations),
                                assigned_sales_locations=sorted(assigned_sales_locations),
                                unassigned_sales_locations=unassigned_sales_locations,
                                assignment_errors=assignment_errors,
                                product_mapping_preview=product_mapping_preview,
                                active_stage=active_stage,
                                countable_products=countable_products,
                                countable_item_options=countable_item_options,
                                countable_selection_errors=countable_selection_errors,
                                gl_codes=_get_purchase_gl_codes(),
                                product_form=product_form,
                                created_product_ids=sorted(created_product_ids),
                                wizard_stage="products",
                            )

                        for product, item_obj in pending_recipe_links:
                            existing_recipe = next(
                                (
                                    recipe
                                    for recipe in product.recipe_items
                                    if recipe.item_id == item_obj.id
                                ),
                                None,
                            )
                            if existing_recipe:
                                existing_recipe.countable = True
                                if not existing_recipe.quantity:
                                    existing_recipe.quantity = 1.0
                            else:
                                db.session.add(
                                    ProductRecipeItem(
                                        product_id=product.id,
                                        item_id=item_obj.id,
                                        quantity=1.0,
                                        countable=True,
                                    )
                                )

                        countable_products = []
                        countable_item_options = []
                        countable_selection_errors = []
                state_data["created_product_ids"] = sorted(created_product_ids)

                if manual_mappings and normalized_lookup:
                    for original_name, product in manual_mappings.items():
                        normalized = normalized_lookup.get(original_name, "")
                        if not normalized:
                            continue
                        alias = alias_lookup.get(normalized)
                        if alias is None:
                            alias = TerminalSaleProductAlias(
                                source_name=original_name,
                                normalized_name=normalized,
                                product=product,
                            )
                            db.session.add(alias)
                            alias_lookup[normalized] = alias
                        else:
                            alias.source_name = original_name
                            alias.product = product

                product_resolution_required = False

            pending_sales: list[dict] = []
            pending_totals: list[dict] = []
            selected_locations: list[str] = []
            issue_queue: list[dict] = []
            preview_lookup: dict[int, dict] = {}
            location_allowed_products: dict[int, set[int]] = {}
            menu_candidate_lookup: dict[int, dict] = {}
            for el in open_locations:
                selected_loc = request.form.get(f"mapping-{el.id}")
                if not selected_loc:
                    continue
                if selected_loc in ignored_sales_locations:
                    continue
                loc_sales = active_sales_summary.get(selected_loc)
                if not loc_sales:
                    continue
                location_updated = False
                price_issues: list[dict] = []
                menu_issues: list[dict] = []
                product_variances: list[dict] = []
                unmatched_entries: list[dict] = []
                price_mismatch_details: list[dict] = []
                menu_issue_details: list[dict] = []
                for prod_name, product_data in loc_sales["products"].items():
                    product = product_lookup.get(prod_name)
                    quantity_value = coerce_float(
                        product_data.get("quantity", 0.0)
                    ) or 0.0
                    file_prices_raw = []
                    file_prices_raw.extend(product_data.get("prices") or [])
                    file_prices_raw.extend(
                        product_data.get("spreadsheet_prices") or []
                    )
                    file_prices = [
                        coerce_float(price)
                        for price in file_prices_raw
                        if price is not None
                    ]
                    file_amount = coerce_float(product_data.get("amount"))
                    if file_amount is None and file_prices:
                        file_amount = quantity_value * file_prices[0]
                    if not product:
                        unmatched_entries.append(
                            {
                                "product_name": prod_name,
                                "quantity": quantity_value,
                                "file_amount": file_amount,
                                "file_prices": file_prices,
                                "sales_location": selected_loc,
                            }
                        )
                        continue
                    allowed_products = location_allowed_products.get(el.id)
                    if allowed_products is None:
                        allowed_products = set()
                        if el.location:
                            allowed_products.update(p.id for p in el.location.products)
                            if el.location.current_menu is not None:
                                allowed_products.update(
                                    p.id for p in el.location.current_menu.products
                                )
                        location_allowed_products[el.id] = allowed_products

                        location_obj = el.location
                        if location_obj and not allowed_products:
                            menu_entry = menu_candidate_lookup.setdefault(
                                el.id,
                                {
                                    "event_location_id": el.id,
                                    "location_name": (
                                        location_obj.name
                                        if location_obj and location_obj.name
                                        else f"Event location #{el.id}"
                                    ),
                                    "menu_name": (
                                        location_obj.current_menu.name
                                        if location_obj and location_obj.current_menu
                                        else None
                                    ),
                                    "products": {},
                                },
                            )
                            menu_entry["products"][product.id] = {
                                "product_id": product.id,
                                "product_name": product.name,
                                "sales_location": selected_loc,
                            }
                            # Menu additions are reviewed separately after the mapping
                            # wizard completes, so defer adding products to the menu here.
                    app_price_value = _terminal_catalog_sell_price(product)
                    pending_sales.append(
                        {
                            "event_location_id": el.id,
                            "product_id": product.id,
                            "product_name": product.name,
                            "source_name": prod_name,
                            "product_price": app_price_value,
                            "normalized_name": normalized_lookup.get(prod_name, ""),
                            "quantity": quantity_value,
                        }
                    )
                    location_updated = True

                    preview_entry = preview_lookup.setdefault(
                        el.id,
                        {
                            "event_location_id": el.id,
                            "event_location_name": (
                                el.location.name
                                if el.location and el.location.name
                                else f"Event location #{el.id}"
                            ),
                            "sales_location_name": selected_loc,
                            "products": [],
                        },
                    )
                    preview_entry["products"].append(
                        {
                            "source_name": prod_name,
                            "matched_product_id": product.id,
                            "matched_product_name": product.name,
                            "quantity": quantity_value,
                            "file_prices": [
                                price for price in file_prices if price is not None
                            ],
                            "file_amount": file_amount,
                            "app_price": app_price_value,
                        }
                    )

                    combined_total_value = combine_terminal_sales_totals(
                        product_data.get("net_including_tax_total"),
                        product_data.get("discount_total"),
                    )
                    derived_unit_price = None
                    if (
                        combined_total_value is not None
                        and quantity_value
                        and abs(quantity_value) > 1e-9
                    ):
                        try:
                            derived_unit_price = float(combined_total_value) / float(
                                quantity_value
                            )
                        except (TypeError, ValueError, ZeroDivisionError):
                            derived_unit_price = None

                    file_price_candidates = [
                        price for price in file_prices if price is not None
                    ]
                    price_candidates: list[float] = []
                    if derived_unit_price is not None:
                        price_candidates.append(derived_unit_price)
                    price_candidates.extend(file_price_candidates)

                    fallback_amount_price = None
                    if file_amount is not None and quantity_value:
                        try:
                            fallback_amount_price = float(file_amount) / float(
                                quantity_value
                            )
                        except (TypeError, ValueError, ZeroDivisionError):
                            fallback_amount_price = None

                    terminal_price_value = derived_unit_price
                    if terminal_price_value is None:
                        if file_price_candidates:
                            terminal_price_value = file_price_candidates[0]
                        else:
                            terminal_price_value = fallback_amount_price

                    if (
                        derived_unit_price is None
                        and not file_price_candidates
                        and fallback_amount_price is not None
                    ):
                        price_candidates.append(fallback_amount_price)

                    if not price_candidates:
                        synthesized_price = None
                        lookup_price = None
                        if product_price_lookup:
                            lookup_price = product_price_lookup.get(prod_name)
                            if lookup_price is None and product is not None:
                                lookup_price = product_price_lookup.get(product.name)
                        lookup_price_value = coerce_float(lookup_price)
                        if lookup_price_value is not None:
                            synthesized_price = lookup_price_value
                        else:
                            location_total_amount = coerce_float(
                                loc_sales.get("total_amount")
                            )
                            location_total_quantity = coerce_float(
                                loc_sales.get("total")
                            )
                            if (
                                location_total_amount is not None
                                and location_total_quantity
                                and abs(location_total_quantity) > 1e-9
                            ):
                                try:
                                    synthesized_price = (
                                        float(location_total_amount)
                                        / float(location_total_quantity)
                                    )
                                except (
                                    TypeError,
                                    ValueError,
                                    ZeroDivisionError,
                                ):
                                    synthesized_price = None

                        if synthesized_price is not None:
                            synthesized_price = float(synthesized_price)
                            price_candidates.append(synthesized_price)
                            terminal_price_value = synthesized_price

                    if not price_candidates and terminal_price_value is not None:
                        price_candidates = [terminal_price_value]

                    matching_candidates = [
                        price
                        for price in price_candidates
                        if _prices_match(price, app_price_value)
                    ]
                    mismatched_candidates = [
                        price
                        for price in price_candidates
                        if not _prices_match(price, app_price_value)
                    ]

                    if mismatched_candidates:
                        catalog_price_value = app_price_value
                        price_mismatch_details.append(
                            {
                                "product_id": product.id,
                                "product_name": product.name,
                                "file_prices": file_prices,
                                "app_price": app_price_value,
                                "catalog_price": catalog_price_value,
                                "terminal_price": terminal_price_value,
                                "sales_location": selected_loc,
                            }
                        )

                        if not matching_candidates:
                            options: dict[str, float | None] = {}
                            if catalog_price_value is not None:
                                options["catalog"] = catalog_price_value
                            if terminal_price_value is not None:
                                options["terminal"] = terminal_price_value

                            price_issues.append(
                                {
                                    "product": product.name,
                                    "product_id": product.id,
                                    "file_prices": file_prices,
                                    "app_price": app_price_value,
                                    "catalog_price": catalog_price_value,
                                    "terminal_price": terminal_price_value,
                                    "sales_location": selected_loc,
                                    "resolution": None,
                                    "selected_price": None,
                                    "selected_option": None,
                                    "target_price": terminal_price_value,
                                    "options": options,
                                }
                            )

                    if (
                        allowed_products
                        and el.location
                        and product.id not in allowed_products
                    ):
                        menu_issues.append(
                            {
                                "product": product.name,
                                "product_id": product.id,
                                "sales_location": selected_loc,
                                "menu_name": (
                                    el.location.current_menu.name
                                    if el.location.current_menu
                                    else None
                                ),
                                "resolution": None,
                            }
                        )
                        menu_issue_details.append(
                            {
                                "product_id": product.id,
                                "product_name": product.name,
                                "sales_location": selected_loc,
                                "menu_name": (
                                    el.location.current_menu.name
                                    if el.location.current_menu
                                    else None
                                ),
                            }
                        )

                    product_variances.append(
                        {
                            "product_id": product.id,
                            "product_name": product.name,
                            "quantity": quantity_value,
                            "file_amount": file_amount,
                            "file_prices": file_prices,
                            "app_price": app_price_value,
                            "sales_location": selected_loc,
                        }
                    )
            variance_details = {
                "products": product_variances,
                "price_mismatches": price_mismatch_details,
                "menu_issues": menu_issue_details,
                "unmapped_products": unmatched_entries,
            }

            if _should_store_terminal_summary(
                loc_sales, location_updated, unmatched_entries
            ):
                selected_locations.append(el.location.name)
                pending_totals.append(
                    {
                        "event_location_id": el.id,
                        "source_location": selected_loc,
                        "total_quantity": loc_sales.get("total"),
                        "total_amount": loc_sales.get("total_amount"),
                        "net_including_tax_total": loc_sales.get(
                            "net_including_tax_total"
                        ),
                        "discount_total": loc_sales.get("discount_total"),
                        "variance_details": variance_details,
                    }
                )

            if location_updated:
                if price_issues or menu_issues:
                    issue_queue.append(
                        {
                            "event_location_id": el.id,
                            "location_name": el.location.name,
                            "sales_location": selected_loc,
                            "price_issues": price_issues,
                            "menu_issues": menu_issues,
                        }
                    )

            if preview_lookup:
                product_mapping_preview = [
                    preview_lookup[el.id]
                    for el in open_locations
                    if el.id in preview_lookup
                ]
            else:
                product_mapping_preview = []

            if menu_candidate_lookup:
                menu_candidates = []
                for el in open_locations:
                    candidate = menu_candidate_lookup.get(el.id)
                    if not candidate:
                        continue
                    products = list(candidate.get("products", {}).values())
                    products.sort(
                        key=lambda info: normalize_name_for_sorting(
                            info.get("product_name", "")
                        )
                    )
                    menu_candidates.append(
                        {
                            "event_location_id": candidate.get("event_location_id", el.id),
                            "location_name": candidate.get("location_name")
                            or (
                                el.location.name
                                if el.location and el.location.name
                                else f"Event location #{el.id}"
                            ),
                            "menu_name": candidate.get("menu_name"),
                            "products": products,
                        }
                    )
            else:
                menu_candidates = []

            existing_selection = (
                state_data.get("menu_candidate_selection")
                if isinstance(state_data, dict)
                else {}
            )
            if not isinstance(existing_selection, dict):
                existing_selection = {}
            selection_map: dict[str, bool] = {}
            for candidate in menu_candidates:
                for product in candidate.get("products", []):
                    key = f"{candidate['event_location_id']}:{product['product_id']}"
                    selection_map[key] = bool(existing_selection.get(key, True))

            if isinstance(state_data, dict):
                state_data["menu_candidates"] = menu_candidates
                if selection_map:
                    state_data["menu_candidate_selection"] = selection_map
                else:
                    state_data.pop("menu_candidate_selection", None)
            elif menu_candidates:
                state_data = {"menu_candidates": menu_candidates}
                if selection_map:
                    state_data["menu_candidate_selection"] = selection_map

            if issue_queue:
                stored_mapping = {
                    str(key): value for key, value in selected_mapping.items()
                }
                state_data = state_data or {}
                state_data["queue"] = issue_queue
                state_data["pending_sales"] = pending_sales
                state_data["selected_locations"] = selected_locations
                state_data["pending_totals"] = pending_totals
                state_data["issue_index"] = 0
                state_data["ignored_sales_locations"] = sorted(ignored_sales_locations)
                state_data["selected_mapping"] = stored_mapping
                state_data["stage"] = "menus"
                state_token, state_data = _save_state(state_data)
                active_stage = "menus"
                return render_template(
                    "events/upload_terminal_sales.html",
                    form=form,
                    event=ev,
                    open_locations=open_locations,
                    mapping_payload=payload,
                    mapping_filename=mapping_filename,
                    sales_summary=sales_summary,
                    sales_location_names=list(sales_summary.keys()),
                    default_mapping=selected_mapping,
                    unresolved_products=[],
                    product_choices=product_choices,
                    product_search_options=product_search_options,
                    skip_selection_value=SKIP_SELECTION_VALUE,
                    create_selection_value=CREATE_SELECTION_VALUE,
                    resolution_errors=resolution_errors,
                    product_resolution_required=False,
                    price_discrepancies={},
                    menu_mismatches={},
                    warnings_required=False,
                    warnings_acknowledged=False,
                    state_token=state_token,
                    issue_index=0,
                    current_issue=issue_queue[0],
                    remaining_locations=len(issue_queue) - 1,
                    selected_locations=selected_locations,
                    issue_total=len(issue_queue),
                    menu_candidates=menu_candidates,
                    menu_candidate_selection=selection_map,
                    ignored_sales_locations=sorted(ignored_sales_locations),
                    assigned_sales_locations=sorted(assigned_sales_locations),
                    unassigned_sales_locations=unassigned_sales_locations,
                    assignment_errors=assignment_errors,
                    product_mapping_preview=product_mapping_preview,
                    active_stage=active_stage,
                    product_form=product_form,
                    created_product_ids=sorted(created_product_ids),
                    wizard_stage="menus",
                )

            if navigate and navigate != "finish":
                active_stage = "locations" if navigate == "back" else "products"
                wizard_stage_value = active_stage
                return render_template(
                    "events/upload_terminal_sales.html",
                    form=form,
                    event=ev,
                    open_locations=open_locations,
                    mapping_payload=payload,
                    mapping_filename=mapping_filename,
                    sales_summary=sales_summary,
                    sales_location_names=list(sales_summary.keys()),
                    default_mapping=selected_mapping,
                    unresolved_products=[],
                    product_choices=product_choices,
                    product_search_options=product_search_options,
                    skip_selection_value=SKIP_SELECTION_VALUE,
                    create_selection_value=CREATE_SELECTION_VALUE,
                    resolution_errors=resolution_errors,
                    product_resolution_required=False,
                    price_discrepancies=price_discrepancies,
                    menu_mismatches=menu_mismatches,
                    warnings_required=warnings_required,
                    warnings_acknowledged=warnings_acknowledged,
                    ignored_sales_locations=sorted(ignored_sales_locations),
                    assigned_sales_locations=sorted(assigned_sales_locations),
                    unassigned_sales_locations=unassigned_sales_locations,
                    assignment_errors=assignment_errors,
                    product_mapping_preview=product_mapping_preview,
                    active_stage=active_stage,
                    countable_products=countable_products,
                    countable_item_options=countable_item_options,
                    countable_selection_errors=countable_selection_errors,
                    gl_codes=_get_purchase_gl_codes() if countable_products else [],
                    product_form=product_form,
                    created_product_ids=sorted(created_product_ids),
                    wizard_stage=wizard_stage_value,
                )

            if navigate == "finish" and menu_candidates:
                state_data = state_data or {}
                state_data.setdefault("queue", [])
                state_data["pending_sales"] = pending_sales
                state_data["pending_totals"] = pending_totals
                state_data["selected_locations"] = selected_locations
                state_data["stage"] = "menus"
                state_token, state_data = _save_state(state_data)
                active_stage = "menus"
                return render_template(
                    "events/upload_terminal_sales.html",
                    form=form,
                    event=ev,
                    open_locations=open_locations,
                    mapping_payload=payload,
                    mapping_filename=mapping_filename,
                    sales_summary=sales_summary,
                    sales_location_names=list(sales_summary.keys()),
                    default_mapping=selected_mapping,
                    unresolved_products=[],
                    product_choices=product_choices,
                    product_search_options=product_search_options,
                    skip_selection_value=SKIP_SELECTION_VALUE,
                    create_selection_value=CREATE_SELECTION_VALUE,
                    resolution_errors=resolution_errors,
                    product_resolution_required=False,
                    price_discrepancies=price_discrepancies,
                    menu_mismatches=menu_mismatches,
                    warnings_required=warnings_required,
                    warnings_acknowledged=warnings_acknowledged,
                    state_token=state_token,
                    issue_index=len(issue_queue),
                    current_issue=None,
                    remaining_locations=0,
                    selected_locations=selected_locations,
                    issue_total=len(issue_queue),
                    menu_candidates=menu_candidates,
                    menu_candidate_selection=selection_map,
                    ignored_sales_locations=sorted(ignored_sales_locations),
                    assigned_sales_locations=sorted(assigned_sales_locations),
                    unassigned_sales_locations=unassigned_sales_locations,
                    assignment_errors=assignment_errors,
                    product_mapping_preview=product_mapping_preview,
                    active_stage=active_stage,
                    countable_products=countable_products,
                    countable_item_options=countable_item_options,
                    countable_selection_errors=countable_selection_errors,
                    gl_codes=_get_purchase_gl_codes() if countable_products else [],
                    product_form=product_form,
                    created_product_ids=sorted(created_product_ids),
                    wizard_stage="menus",
                )

            updated_locations = _apply_pending_sales(
                pending_sales,
                pending_totals,
                link_products_to_locations=True,
            )
            saved_location_aliases = _store_location_aliases(pending_totals)
            if updated_locations or saved_location_aliases:
                db.session.commit()
                log_activity(
                    "Uploaded terminal sales for event "
                    f"{event_id} from {mapping_filename or 'uploaded file'}"
                )
                success_parts = []
                if updated_locations:
                    success_parts.append(
                        "Terminal sales were imported for: "
                        + ", ".join(sorted(updated_locations))
                    )
                if saved_location_aliases:
                    success_parts.append(
                        "Remembered location mappings for: "
                        + ", ".join(sorted(saved_location_aliases))
                    )
                flash(" ".join(success_parts), "success")
            else:
                flash(
                    "No event locations were linked to the uploaded sales data.",
                    "warning",
                )
            _clear_state()
            return redirect(url_for("event.view_event", event_id=event_id))
        elif step:
            flash("Unable to process the uploaded sales data.", "danger")
            return redirect(url_for("event.upload_terminal_sales", event_id=event_id))

    if form.validate_on_submit():
        file = form.file.data
        filename = secure_filename(file.filename)
        ext = os.path.splitext(filename)[1].lower()
        filepath = os.path.join(current_app.config["UPLOAD_FOLDER"], filename)
        file.save(filepath)

        rows: list[dict] = []

        def add_row(
            loc,
            name,
            qty,
            price=None,
            raw_price=None,
            amount=None,
            net_including_tax_total=None,
            discounts_total=None,
            *,
            is_location_total: bool = False,
        ):
            if not loc or not isinstance(loc, str):
                return
            loc = loc.strip()
            if not loc:
                return
            product_name = None
            if not is_location_total:
                if not name or not isinstance(name, str):
                    return
                product_name = name.strip()
                if not product_name:
                    return
            elif isinstance(name, str):
                product_name = name.strip() or None
            price_value = parse_terminal_sales_number(price)
            raw_price_value = parse_terminal_sales_number(raw_price)

            if price_value is None and raw_price_value is not None:
                price_value = raw_price_value
                raw_price_value = None

            amount_value = parse_terminal_sales_number(amount)
            net_including_value = parse_terminal_sales_number(
                net_including_tax_total
            )
            discounts_value = parse_terminal_sales_number(discounts_total)

            quantity_value = parse_terminal_sales_number(qty)
            quantity_value = derive_terminal_sales_quantity(
                quantity_value,
                price=price_value,
                amount=amount_value,
                net_including_tax_total=net_including_value,
                discounts_total=discounts_value,
            )
            if quantity_value is None and not is_location_total:
                return
            entry = {
                "location": loc,
            }
            if is_location_total:
                entry["is_location_total"] = True
            if product_name:
                entry["product"] = product_name
            if quantity_value is not None:
                entry["quantity"] = quantity_value
            if price_value is not None:
                entry["price"] = price_value
            if raw_price_value is not None:
                entry["raw_price"] = raw_price_value
            if amount_value is not None:
                entry["amount"] = amount_value
            if net_including_value is not None:
                entry["net_including_tax_total"] = net_including_value
            if discounts_value is not None:
                entry["discount_total"] = discounts_value
            rows.append(entry)

        try:
            if ext in {".xls", ".xlsx"}:
                def _iter_excel_rows(path: str, extension: str):
                    if extension == ".xls":
                        try:
                            import xlrd  # type: ignore
                        except ModuleNotFoundError:
                            try:
                                from app.vendor import xlrd  # type: ignore
                            except ImportError:
                                raise RuntimeError("legacy_xls_missing") from None
                        try:
                            book = xlrd.open_workbook(path)
                        except Exception as exc:  # pragma: no cover - defensive
                            raise RuntimeError("legacy_xls_error") from exc

                        try:
                            sheet = book.sheet_by_index(0)
                        except IndexError as exc:
                            raise RuntimeError("legacy_xls_error") from exc

                        try:
                            for row_idx in range(sheet.nrows):
                                yield [
                                    sheet.cell_value(row_idx, col_idx)
                                    for col_idx in range(sheet.ncols)
                                ]
                        finally:  # pragma: no branch - ensure resources freed
                            try:
                                book.release_resources()
                            except AttributeError:
                                pass
                    else:
                        try:
                            from openpyxl import load_workbook
                        except ImportError as exc:  # pragma: no cover - env issue
                            raise RuntimeError("xlsx_missing") from exc

                        try:
                            workbook = load_workbook(
                                path, read_only=True, data_only=True
                            )
                        except Exception as exc:
                            raise RuntimeError("xlsx_error") from exc

                        try:
                            sheet = workbook.active
                            for row in sheet.iter_rows(values_only=True):
                                yield list(row)
                        finally:
                            workbook.close()

                try:
                    rows_iter = _iter_excel_rows(filepath, ext)
                except RuntimeError as exc:
                    reason = str(exc)
                    if reason == "legacy_xls_missing":
                        flash(
                            "Legacy Excel support is unavailable on this server.",
                            "danger",
                        )
                        current_app.logger.exception(
                            "xlrd is required to read legacy .xls files"
                        )
                    else:
                        flash(
                            "The uploaded Excel file could not be read.",
                            "danger",
                        )
                        current_app.logger.exception(
                            "Failed to parse Excel file during terminal sales upload"
                        )
                    return redirect(
                        url_for("event.upload_terminal_sales", event_id=event_id)
                    )

                current_loc = None
                for row in rows_iter:
                    location_name = extract_terminal_sales_location(row)
                    if location_name:
                        current_loc = location_name
                        continue

                    if not current_loc:
                        continue

                    second = row[1] if len(row) > 1 else None
                    first_cell = row[0] if row else None
                    quantity_cell = row[4] if len(row) > 4 else None
                    amount_cell = row[5] if len(row) > 5 else None
                    net_cell = row[7] if len(row) > 7 else None
                    discount_cell = row[8] if len(row) > 8 else None

                    summary_quantity = parse_terminal_sales_number(quantity_cell)
                    summary_amount = parse_terminal_sales_number(amount_cell)
                    summary_net = parse_terminal_sales_number(net_cell)
                    summary_discount = parse_terminal_sales_number(discount_cell)

                    if (
                        terminal_sales_cell_is_blank(first_cell)
                        and not isinstance(second, str)
                        and (
                            summary_quantity is not None
                            or summary_amount is not None
                            or summary_net is not None
                            or summary_discount is not None
                        )
                    ):
                        add_row(
                            current_loc,
                            None,
                            quantity_cell,
                            price=None,
                            amount=amount_cell,
                            net_including_tax_total=net_cell,
                            discounts_total=discount_cell,
                            is_location_total=True,
                        )
                        continue

                    if not isinstance(second, str):
                        continue

                    quantity = quantity_cell
                    price_cell = row[2] if len(row) > 2 else None
                    quantity_value = summary_quantity
                    discounts = summary_discount
                    combined_total_value = combine_terminal_sales_totals(
                        summary_net, summary_discount
                    )
                    computed_price = None
                    if (
                        combined_total_value is not None
                        and quantity_value is not None
                        and abs(quantity_value) > 1e-9
                    ):
                        try:
                            computed_price = float(combined_total_value) / float(
                                quantity_value
                            )
                        except (TypeError, ValueError, ZeroDivisionError):
                            computed_price = None
                    price = computed_price if computed_price is not None else price_cell
                    raw_price_cell = price_cell
                    amount = amount_cell
                    net_including_total = net_cell
                    add_row(
                        current_loc,
                        second,
                        quantity,
                        price,
                        raw_price=raw_price_cell,
                        amount=amount,
                        net_including_tax_total=net_including_total,
                        discounts_total=discounts,
                    )
            elif ext == ".pdf":
                import pdfplumber

                try:
                    with pdfplumber.open(filepath) as pdf:
                        text = "\n".join(
                            page.extract_text() or "" for page in pdf.pages
                        )
                except Exception:
                    current_app.logger.exception(
                        "Failed to parse PDF file during terminal sales upload"
                    )
                    flash(
                        "The uploaded PDF file could not be read. "
                        "Please upload a valid sales export.",
                        "danger",
                    )
                    return redirect(
                        url_for("event.upload_terminal_sales", event_id=event_id)
                    )
                current_loc = None
                for line in text.splitlines():
                    line = line.strip()
                    if not line:
                        continue
                    if not line[0].isdigit():
                        current_loc = line
                        continue
                    if current_loc is None:
                        continue
                    parts = line.split()
                    idx = 1
                    while (
                        idx < len(parts)
                        and not parts[idx].replace(".", "", 1).isdigit()
                    ):
                        idx += 1
                    if idx + 2 < len(parts):
                        name = " ".join(parts[1:idx])
                        qty = parts[idx + 2]
                        add_row(current_loc, name, qty)
        finally:
            try:
                os.remove(filepath)
            except OSError:
                pass

        if not rows:
            flash(
                "No sales records were detected in the uploaded file.",
                "warning",
            )
        else:
            rows_data = rows

            sales_summary = _group_rows(rows_data)
            sales_location_names = list(sales_summary.keys())
            mapping_payload = json.dumps(
                {"rows": rows_data, "filename": filename}
            )
            mapping_filename = filename
            default_mapping = suggest_terminal_sales_location_mapping(
                open_locations, sales_summary
            )
            assigned_sales_locations = {
                value for value in default_mapping.values() if value
            }
            unassigned_sales_locations = [
                name
                for name in sales_location_names
                if name not in assigned_sales_locations
            ]

            _clear_state()
            initial_state = {
                "stage": "locations",
                "payload": {"rows": rows_data, "filename": filename},
                "mapping_filename": filename,
                "selected_mapping": {
                    str(key): value for key, value in default_mapping.items()
                },
                "ignored_sales_locations": [],
                "warnings_acknowledged": False,
                "created_product_ids": [],
                "product_creations": {},
            }
            state_token, state_data = _save_state(initial_state)

    if state_data:
        wizard_stage = state_data.get("stage", wizard_stage)
        issue_index = state_data.get("issue_index", issue_index)
        selected_locations = state_data.get("selected_locations") or selected_locations
        created_product_ids_state = set(
            state_data.get("created_product_ids") or []
        )
        product_creations_state = dict(state_data.get("product_creations") or {})

    if request.method != "POST" and state_data and wizard_stage in {"locations", "products", "menus"}:
        payload_data = state_data.get("payload") or {}
        if payload_data:
            mapping_payload = json.dumps(payload_data)
            rows = payload_data.get("rows") or []
            sales_summary = _group_rows(rows)
            sales_location_names = list(sales_summary.keys())
        mapping_filename = state_data.get("mapping_filename") or mapping_filename
        stored_mapping = state_data.get("selected_mapping") or {}
        default_mapping = {}
        for key, value in stored_mapping.items():
            try:
                default_mapping[int(key)] = value
            except (TypeError, ValueError):
                continue
        ignored_sales_locations = set(state_data.get("ignored_sales_locations") or [])
        assigned_sales_locations = {value for value in default_mapping.values() if value}
        unassigned_sales_locations = [
            name
            for name in sales_location_names
            if name not in assigned_sales_locations
            and name not in ignored_sales_locations
        ]
        if wizard_stage == "products":
            product_resolution_required = True
            if not product_choices:
                product_choices = Product.query.order_by(Product.name).all()
            product_search_options = [
                {
                    "id": str(product.id),
                    "value": f"{product.name} (ID: {product.id})",
                    "label": product.name,
                }
                for product in product_choices
            ]
            active_sales_summary = {
                name: data
                for name, data in sales_summary.items()
                if name not in ignored_sales_locations
            }
            product_price_lookup = _derive_price_map(active_sales_summary)
            product_selections = state_data.get("product_selections") or {}
            unresolved_products = []
            for idx, original_name in enumerate(product_selections.keys()):
                unresolved_products.append(
                    {
                        "field": f"product-match-{idx}",
                        "name": original_name,
                        "selected": product_selections.get(original_name, ""),
                        "price": product_price_lookup.get(original_name),
                        "created_product_id": product_creations_state.get(
                            original_name
                        ),
                    }
                )
            active_stage = "products"
        elif wizard_stage == "menus":
            queue = state_data.get("queue") or []
            total_locations = len(queue)
            if queue and 0 <= issue_index < len(queue):
                current_issue = queue[issue_index]
                remaining_locations = len(queue) - issue_index - 1
            selected_locations = state_data.get("selected_locations") or []
            warnings_acknowledged = state_data.get("warnings_acknowledged", False)
            active_stage = "menus"
        else:
            active_stage = "locations"

    return render_template(
        "events/upload_terminal_sales.html",
        form=form,
        event=ev,
        open_locations=open_locations,
        mapping_payload=mapping_payload,
        mapping_filename=mapping_filename,
        sales_summary=sales_summary,
        sales_location_names=sales_location_names,
        default_mapping=default_mapping,
        unresolved_products=unresolved_products,
        product_choices=product_choices,
        product_search_options=product_search_options,
        skip_selection_value=SKIP_SELECTION_VALUE,
        create_selection_value=CREATE_SELECTION_VALUE,
        resolution_errors=resolution_errors,
        product_resolution_required=product_resolution_required,
        price_discrepancies=price_discrepancies,
        menu_mismatches=menu_mismatches,
        warnings_required=warnings_required,
        warnings_acknowledged=warnings_acknowledged,
        state_token=state_token,
        current_issue=current_issue,
        issue_index=issue_index,
        remaining_locations=remaining_locations,
        selected_locations=selected_locations,
        issue_total=total_locations,
        ignored_sales_locations=sorted(ignored_sales_locations),
        assigned_sales_locations=sorted(assigned_sales_locations),
        unassigned_sales_locations=unassigned_sales_locations,
        assignment_errors=assignment_errors,
        product_mapping_preview=product_mapping_preview,
        active_stage=active_stage,
        wizard_stage=wizard_stage,
        countable_products=countable_products,
        countable_item_options=countable_item_options,
        countable_selection_errors=countable_selection_errors,
        gl_codes=_get_purchase_gl_codes() if countable_products else [],
        product_form=product_form,
        created_product_ids=sorted(created_product_ids_state),
    )


@event.route(
    "/events/<int:event_id>/locations/<int:el_id>/confirm",
    methods=["GET", "POST"],
)
@login_required
def confirm_location(event_id, el_id):
    el = db.session.get(EventLocation, el_id)
    if el is None or el.event_id != event_id:
        abort(404)
    form = EventLocationConfirmForm()
    if form.validate_on_submit():
        summary_record = el.terminal_sales_summary
        manual_sales = (
            TerminalSale.query.filter_by(event_location_id=el.id).all()
        )
        has_file_summary = False
        if summary_record is not None:
            has_file_summary = any(
                field is not None
                for field in (
                    summary_record.total_amount,
                    summary_record.total_quantity,
                )
            ) or bool(summary_record.variance_details)

        if manual_sales and not has_file_summary:
            if summary_record is None:
                summary_record = EventLocationTerminalSalesSummary(
                    event_location=el
                )
                db.session.add(summary_record)

            total_quantity = sum(float(sale.quantity or 0.0) for sale in manual_sales)
            total_amount = sum(
                float(sale.quantity or 0.0)
                * float(getattr(sale.product, "price", 0.0) or 0.0)
                for sale in manual_sales
            )
            summary_record.total_quantity = total_quantity
            summary_record.total_amount = total_amount

        el.confirmed = True
        db.session.commit()
        log_activity(
            f"Confirmed event location {el_id} for event {event_id}"
        )
        flash("Location confirmed")
        return redirect(url_for("event.view_event", event_id=event_id))
    location, stand_items = _get_stand_items(el.location_id, event_id)
    stand_sheet_item_ids: set[int] = set()
    if location is not None:
        stand_sheet_item_ids.update(
            item.item_id for item in (location.stand_items or []) if item.item_id
        )

    stand_sheet_item_ids.update(
        sheet.item_id for sheet in (el.stand_sheet_items or []) if sheet.item_id
    )

    untracked_sales: list[dict[str, object]] = []
    if el.terminal_sales:
        aggregated: dict[int, dict[str, object]] = {}
        for sale in el.terminal_sales:
            quantity = float(sale.quantity or 0.0)
            if not quantity:
                continue

            product = sale.product
            if product is None:
                continue

            countable_items = [
                ri.item_id
                for ri in (product.recipe_items or [])
                if ri.countable and ri.item_id is not None
            ]
            if not countable_items:
                continue

            if not all(item_id in stand_sheet_item_ids for item_id in countable_items):
                amount = quantity * float(getattr(product, "price", 0.0) or 0.0)
                product_id = product.id
                if product_id in aggregated:
                    aggregated_entry = aggregated[product_id]
                    aggregated_entry["quantity"] += quantity
                    aggregated_entry["amount"] += amount
                else:
                    aggregated[product_id] = {
                        "product": product,
                        "product_name": getattr(product, "name", "Unnamed product"),
                        "quantity": quantity,
                        "amount": amount,
                    }

        if aggregated:
            untracked_sales = sorted(
                aggregated.values(),
                key=lambda entry: entry["product_name"].casefold()
                if isinstance(entry.get("product_name"), str)
                else "",
            )
    stand_variances: list[dict] = []
    conversions = _conversion_mapping()
    price_lookup = _build_item_price_lookup(el, stand_items)
    for entry in stand_items:
        sheet_values = entry.get("sheet_values")
        opening_val = getattr(sheet_values, "opening_count", None) or 0.0
        in_val = getattr(sheet_values, "transferred_in", None) or 0.0
        out_val = getattr(sheet_values, "transferred_out", None) or 0.0
        eaten_val = getattr(sheet_values, "eaten", None) or 0.0
        spoiled_val = getattr(sheet_values, "spoiled", None) or 0.0
        adjustments_val = getattr(sheet_values, "adjustments", None) or 0.0
        closing_val = getattr(sheet_values, "closing_count", None) or 0.0
        sales_val = entry.get("sales") or 0.0
        variance = (
            opening_val
            + in_val
            + adjustments_val
            - out_val
            - sales_val
            - eaten_val
            - spoiled_val
            - closing_val
        )
        has_sheet = entry.get("sheet") is not None
        item_obj = entry.get("item")
        price_per_unit_base = (
            price_lookup.get(item_obj.id)
            if item_obj is not None and price_lookup is not None
            else None
        )
        price_per_unit = (
            convert_cost_for_reporting(
                price_per_unit_base, entry.get("base_unit"), conversions
            )
            if price_per_unit_base is not None
            else None
        )
        variance_amount = (
            variance * price_per_unit
            if has_sheet and price_per_unit is not None
            else None
        )
        stand_variances.append(
            {
                "item": entry.get("item"),
                "report_unit_label": entry.get("report_unit_label"),
                "variance": variance if has_sheet else 0.0,
                "closing": closing_val if has_sheet else None,
                "price": price_per_unit,
                "variance_amount": variance_amount,
            }
        )

    app_total_quantity = sum(float(sale.quantity or 0.0) for sale in el.terminal_sales)
    app_total_amount = sum(
        float(sale.quantity or 0.0) * float(sale.product.price or 0.0)
        for sale in el.terminal_sales
    )
    summary_record = el.terminal_sales_summary
    file_total_amount = None
    file_total_quantity = None
    source_location_name = None
    if summary_record is not None:
        file_total_amount = summary_record.total_amount
        file_total_quantity = summary_record.total_quantity
        source_location_name = summary_record.source_location
        fallback_quantity, fallback_amount = _derive_summary_totals_from_details(
            summary_record.variance_details
        )
        if file_total_quantity is None and fallback_quantity is not None:
            file_total_quantity = fallback_quantity
        if file_total_amount is None and fallback_amount is not None:
            file_total_amount = fallback_amount
    amount_variance = None
    if file_total_amount is not None:
        amount_variance = app_total_amount - float(file_total_amount)

    variance_breakdown = {
        "product_deltas": [],
        "price_mismatches": [],
        "menu_issues": [],
        "unmapped_products": [],
        "summary": [],
        "has_details": False,
        "show_toggle": False,
    }
    if summary_record is not None and summary_record.variance_details:
        details = _normalize_variance_details(summary_record.variance_details) or {}
        product_entries = details.get("products") or []
        price_entries = details.get("price_mismatches") or []
        menu_entries = details.get("menu_issues") or []

        product_ids = {
            entry.get("product_id")
            for entry in (product_entries + price_entries + menu_entries)
        }
        product_ids.discard(None)
        resolved_products: dict[int, Product] = {}
        if product_ids:
            resolved_products = {
                p.id: p
                for p in Product.query.filter(Product.id.in_(product_ids)).all()
            }

        product_deltas: list[dict] = []
        for entry in product_entries:
            product_obj = None
            product_id = entry.get("product_id")
            if product_id in resolved_products:
                product_obj = resolved_products[product_id]
            name = (
                product_obj.name
                if product_obj is not None
                else entry.get("product_name")
            ) or "Unknown product"
            quantity = coerce_float(entry.get("quantity")) or 0.0
            app_price = (
                coerce_float(product_obj.price)
                if product_obj is not None
                else coerce_float(entry.get("app_price"))
            )
            app_amount = quantity * app_price if app_price is not None else None
            file_amount = coerce_float(entry.get("file_amount"))
            file_prices = [
                price
                for price in (entry.get("file_prices") or [])
                if price is not None
            ]
            amount_delta = None
            if app_amount is not None and file_amount is not None:
                amount_delta = app_amount - file_amount
            product_deltas.append(
                {
                    "product_name": name,
                    "quantity": quantity,
                    "app_price": app_price,
                    "file_prices": file_prices,
                    "app_amount": app_amount,
                    "file_amount": file_amount,
                    "amount_delta": amount_delta,
                }
            )

        price_mismatches: list[dict] = []
        for entry in price_entries:
            product_obj = None
            product_id = entry.get("product_id")
            if product_id in resolved_products:
                product_obj = resolved_products[product_id]
            name = (
                product_obj.name
                if product_obj is not None
                else entry.get("product_name")
            ) or "Unknown product"
            app_price = (
                coerce_float(product_obj.price)
                if product_obj is not None
                else coerce_float(entry.get("app_price"))
            )
            file_prices = [
                price
                for price in (entry.get("file_prices") or [])
                if price is not None
            ]
            price_mismatches.append(
                {
                    "product_name": name,
                    "app_price": app_price,
                    "file_prices": file_prices,
                    "sales_location": entry.get("sales_location"),
                }
            )

        menu_issue_details: list[dict] = []
        for entry in menu_entries:
            product_obj = None
            product_id = entry.get("product_id")
            if product_id in resolved_products:
                product_obj = resolved_products[product_id]
            name = (
                product_obj.name
                if product_obj is not None
                else entry.get("product_name")
            ) or "Unknown product"
            menu_issue_details.append(
                {
                    "product_name": name,
                    "menu_name": entry.get("menu_name"),
                    "sales_location": entry.get("sales_location"),
                }
            )

        unmatched_details: list[dict] = []
        for entry in details.get("unmapped_products", []) or []:
            file_prices = [
                price
                for price in (entry.get("file_prices") or [])
                if price is not None
            ]
            unmatched_details.append(
                {
                    "product_name": entry.get("product_name") or "Unmapped entry",
                    "quantity": coerce_float(entry.get("quantity")),
                    "file_amount": coerce_float(entry.get("file_amount")),
                    "file_prices": file_prices,
                    "sales_location": entry.get("sales_location"),
                }
            )

        variance_breakdown["product_deltas"] = product_deltas
        variance_breakdown["price_mismatches"] = price_mismatches
        variance_breakdown["menu_issues"] = menu_issue_details
        variance_breakdown["unmapped_products"] = unmatched_details

        summary_items: list[dict[str, object]] = []
        if product_deltas:
            delta_values = [
                entry["amount_delta"]
                for entry in product_deltas
                if entry.get("amount_delta") is not None
            ]
            summary_items.append(
                {
                    "label": "Product differences",
                    "count": len(product_deltas),
                    "impact": sum(delta_values) if delta_values else None,
                }
            )
        if price_mismatches:
            summary_items.append(
                {
                    "label": "Price mismatches",
                    "count": len(price_mismatches),
                    "impact": None,
                }
            )
        if menu_issue_details:
            summary_items.append(
                {
                    "label": "Menu issues",
                    "count": len(menu_issue_details),
                    "impact": None,
                }
            )
        if unmatched_details:
            unmatched_amounts = [
                entry["file_amount"]
                for entry in unmatched_details
                if entry.get("file_amount") is not None
            ]
            summary_items.append(
                {
                    "label": "Unmapped or missing items",
                    "count": len(unmatched_details),
                    "impact": sum(unmatched_amounts) if unmatched_amounts else None,
                }
            )

        variance_breakdown["summary"] = summary_items
        variance_breakdown["has_details"] = any(
            (
                product_deltas,
                price_mismatches,
                menu_issue_details,
                unmatched_details,
            )
        )

    variance_breakdown["show_toggle"] = bool(
        variance_breakdown["has_details"]
        or (amount_variance is not None)
    )
    return render_template(
        "events/confirm_location.html",
        form=form,
        event_location=el,
        stand_variances=stand_variances,
        sales_summary={
            "app_total_quantity": app_total_quantity,
            "app_total_amount": app_total_amount,
            "file_total_quantity": file_total_quantity,
            "file_total_amount": file_total_amount,
            "amount_variance": amount_variance,
            "source_location": source_location_name,
        },
        variance_breakdown=variance_breakdown,
        location=location,
        untracked_sales=untracked_sales,
    )


@event.route(
    "/events/<int:event_id>/locations/<int:el_id>/undo-confirm",
    methods=["POST"],
)
@login_required
def undo_confirm_location(event_id, el_id):
    el = db.session.get(EventLocation, el_id)
    if el is None or el.event_id != event_id:
        abort(404)

    form = EventLocationUndoConfirmForm()
    if not form.validate_on_submit():
        flash("Unable to undo the confirmation. Please try again.", "warning")
        return redirect(url_for("event.view_event", event_id=event_id))

    ev = el.event
    if ev is not None and ev.closed:
        flash(
            "This event is closed and location confirmations cannot be changed.",
            "warning",
        )
        return redirect(url_for("event.view_event", event_id=event_id))

    if not el.confirmed:
        flash("This location has not been confirmed.", "warning")
        return redirect(url_for("event.view_event", event_id=event_id))

    el.confirmed = False
    db.session.commit()

    log_activity(
        f"Reopened event location {el_id} for event {event_id}"
    )
    flash("Location confirmation undone.", "success")

    return redirect(url_for("event.view_event", event_id=event_id))


def _get_stand_items(location_id, event_id=None):
    location = db.session.get(Location, location_id)
    conversions = _conversion_mapping()
    stand_items = []
    seen = set()

    sales_by_item = {}
    sheet_map = {}
    if event_id is not None:
        el = EventLocation.query.filter_by(
            event_id=event_id,
            location_id=location_id,
        ).first()
        if el:
            for sale in el.terminal_sales:
                for ri in sale.product.recipe_items:
                    if ri.countable:
                        factor = ri.unit.factor if ri.unit else 1
                        sales_by_item[ri.item_id] = (
                            sales_by_item.get(ri.item_id, 0)
                            + sale.quantity * ri.quantity * factor
                        )
            for sheet in el.stand_sheet_items:
                sheet_map[sheet.item_id] = sheet

    authoritative_products = get_authoritative_location_products(location)
    drift_item_ids = get_location_drift_recipe_item_ids(location)

    for product_obj in authoritative_products:
        for recipe_item in product_obj.recipe_items:
            if recipe_item.countable and recipe_item.item_id not in seen:
                seen.add(recipe_item.item_id)
                record = LocationStandItem.query.filter_by(
                    location_id=location_id,
                    item_id=recipe_item.item_id,
                ).first()
                expected = record.expected_count if record else 0
                sales = sales_by_item.get(recipe_item.item_id, 0)
                item = recipe_item.item
                recv_unit = next(
                    (u for u in item.units if u.receiving_default), None
                )
                trans_unit = next(
                    (u for u in item.units if u.transfer_default), None
                )
                stand_items.append(
                    _build_stand_item_entry(
                        item=item,
                        expected=expected,
                        sales=sales,
                        sheet=sheet_map.get(recipe_item.item_id),
                        recv_unit=recv_unit,
                        trans_unit=trans_unit,
                        conversions=conversions,
                    )
                )

    # Include any items directly assigned to the location that may not be
    # part of a product recipe (e.g. items received via purchase invoices).
    for record in LocationStandItem.query.filter_by(
        location_id=location_id
    ).all():
        if record.item_id in seen:
            continue
        if record.item_id in drift_item_ids:
            continue
        item = record.item
        recv_unit = next((u for u in item.units if u.receiving_default), None)
        trans_unit = next((u for u in item.units if u.transfer_default), None)
        stand_items.append(
            _build_stand_item_entry(
                item=item,
                expected=record.expected_count,
                sales=sales_by_item.get(record.item_id, 0),
                sheet=sheet_map.get(record.item_id),
                recv_unit=recv_unit,
                trans_unit=trans_unit,
                conversions=conversions,
            )
        )
        seen.add(record.item_id)

    stand_items.sort(
        key=lambda entry: normalize_name_for_sorting(
            entry["item"].name
        ).casefold(),
    )

    return location, stand_items


def build_sustainability_report(event_id: int) -> dict:
    """Aggregate waste, cost, and carbon metrics for an event."""

    carbon_per_unit = float(
        current_app.config.get("CARBON_EQ_PER_UNIT", 0.5)
    )
    query = (
        db.session.query(
            EventStandSheetItem,
            EventLocation,
            Location,
            Item,
        )
        .join(EventStandSheetItem.event_location)
        .join(EventLocation.location)
        .join(EventStandSheetItem.item)
        .filter(EventLocation.event_id == event_id)
    )

    totals = {"waste": 0.0, "cost": 0.0, "carbon": 0.0}
    location_totals = defaultdict(
        lambda: {"waste": 0.0, "cost": 0.0, "carbon": 0.0}
    )
    item_totals = defaultdict(lambda: {"waste": 0.0, "cost": 0.0, "carbon": 0.0})

    for sheet, _, location, item in query.all():
        eaten = sheet.eaten or 0.0
        spoiled = sheet.spoiled or 0.0
        waste_units = eaten + spoiled
        if waste_units == 0:
            continue

        unit_cost = item.cost or 0.0
        carbon_factor = getattr(item, "carbon_factor", None)
        if carbon_factor is None:
            carbon_factor = carbon_per_unit

        waste_cost = waste_units * unit_cost
        carbon_eq = waste_units * carbon_factor

        totals["waste"] += waste_units
        totals["cost"] += waste_cost
        totals["carbon"] += carbon_eq

        loc_bucket = location_totals[location.name]
        loc_bucket["waste"] += waste_units
        loc_bucket["cost"] += waste_cost
        loc_bucket["carbon"] += carbon_eq

        item_bucket = item_totals[item.name]
        item_bucket["waste"] += waste_units
        item_bucket["cost"] += waste_cost
        item_bucket["carbon"] += carbon_eq

    location_breakdown = [
        {
            "location": name,
            "waste": values["waste"],
            "cost": values["cost"],
            "carbon": values["carbon"],
        }
        for name, values in sorted(
            location_totals.items(), key=lambda item: item[1]["waste"], reverse=True
        )
    ]

    item_leaderboard = [
        {
            "item": name,
            "waste": values["waste"],
            "cost": values["cost"],
            "carbon": values["carbon"],
        }
        for name, values in sorted(
            item_totals.items(), key=lambda item: item[1]["waste"], reverse=True
        )
    ]

    goal_target = float(
        current_app.config.get("SUSTAINABILITY_WASTE_GOAL", 0) or 0.0
    )
    goal_progress = None
    goal_remaining = None
    goal_met = None
    if goal_target > 0:
        goal_remaining = max(goal_target - totals["waste"], 0.0)
        consumed_pct = min((totals["waste"] / goal_target) * 100, 100)
        goal_progress = round(100 - consumed_pct, 2)
        goal_met = totals["waste"] <= goal_target

    chart_data = {
        "labels": [entry["location"] for entry in location_breakdown],
        "datasets": [
            {
                "label": "Waste (units)",
                "backgroundColor": "#198754",
                "data": [entry["waste"] for entry in location_breakdown],
            },
            {
                "label": "Carbon (kg CO₂e)",
                "backgroundColor": "#0d6efd",
                "data": [entry["carbon"] for entry in location_breakdown],
            },
        ],
    }

    return {
        "totals": totals,
        "location_breakdown": location_breakdown,
        "item_leaderboard": item_leaderboard,
        "goal": {
            "target": goal_target if goal_target > 0 else None,
            "remaining": goal_remaining,
            "progress_pct": goal_progress,
            "met": goal_met,
        },
        "chart_data": chart_data,
    }


@event.route(
    "/events/<int:event_id>/stand_sheet/<int:location_id>",
    methods=["GET", "POST"],
)
@login_required
def stand_sheet(event_id, location_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)
    el = EventLocation.query.filter_by(
        event_id=event_id, location_id=location_id
    ).first()
    if el is None:
        abort(404)
    if el.confirmed or ev.closed:
        flash(
            "This location is closed and the stand sheet cannot be modified."
        )
        return redirect(url_for("event.view_event", event_id=event_id))

    location, stand_items = _get_stand_items(location_id, event_id)

    if request.method == "POST":
        for entry in stand_items:
            item_id = entry["item"].id
            base_unit = entry.get("base_unit")
            report_unit = entry.get("report_unit") or base_unit
            sheet = EventStandSheetItem.query.filter_by(
                event_location_id=el.id,
                item_id=item_id,
            ).first()
            if not sheet:
                sheet = EventStandSheetItem(
                    event_location_id=el.id, item_id=item_id
                )
                db.session.add(sheet)
            opening = coerce_float(request.form.get(f"open_{item_id}"), default=0.0)
            transferred_in = coerce_float(
                request.form.get(f"in_{item_id}"), default=0.0
            )
            transferred_out = coerce_float(
                request.form.get(f"out_{item_id}"), default=0.0
            )
            adjustments = coerce_float(
                request.form.get(f"adjust_{item_id}"), default=0.0
            )
            eaten = coerce_float(request.form.get(f"eaten_{item_id}"), default=0.0)
            spoiled = coerce_float(
                request.form.get(f"spoiled_{item_id}"), default=0.0
            )
            closing = coerce_float(request.form.get(f"close_{item_id}"), default=0.0)
            sheet.opening_count = _convert_report_value_to_base(
                opening or 0, base_unit, report_unit
            )
            sheet.transferred_in = _convert_report_value_to_base(
                transferred_in or 0, base_unit, report_unit
            )
            sheet.transferred_out = _convert_report_value_to_base(
                transferred_out or 0, base_unit, report_unit
            )
            sheet.adjustments = _convert_report_value_to_base(
                adjustments or 0, base_unit, report_unit
            )
            sheet.eaten = _convert_report_value_to_base(
                eaten or 0, base_unit, report_unit
            )
            sheet.spoiled = _convert_report_value_to_base(
                spoiled or 0, base_unit, report_unit
            )
            sheet.closing_count = _convert_report_value_to_base(
                closing or 0, base_unit, report_unit
            )
        notes_text = request.form.get("notes", "")
        el.notes = notes_text.strip() or None
        db.session.commit()
        log_activity(
            f"Updated stand sheet for event {event_id} location {location_id}"
        )
        flash("Stand sheet saved")
        return redirect(url_for("event.view_event", event_id=event_id))

    return render_template(
        "events/stand_sheet.html",
        event=ev,
        location=location,
        stand_items=stand_items,
        event_location=el,
    )


@event.route("/events/<int:event_id>/sustainability")
@login_required
def sustainability_dashboard(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)

    report = build_sustainability_report(event_id)
    return render_template(
        "events/sustainability_dashboard.html",
        event=ev,
        report=report,
        chart_data=report["chart_data"],
        print_view=False,
    )


@event.route("/events/<int:event_id>/sustainability/print")
@login_required
def sustainability_dashboard_print(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)

    report = build_sustainability_report(event_id)
    return render_template(
        "events/sustainability_dashboard.html",
        event=ev,
        report=report,
        chart_data=report["chart_data"],
        print_view=True,
    )


@event.route("/events/<int:event_id>/sustainability/export.csv")
@login_required
def sustainability_dashboard_csv(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)

    report = build_sustainability_report(event_id)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Location", "Waste (units)", "Cost", "Carbon (kg CO₂e)"])
    for entry in report["location_breakdown"]:
        writer.writerow(
            [
                entry["location"],
                f"{entry['waste']:.2f}",
                f"{entry['cost']:.2f}",
                f"{entry['carbon']:.2f}",
            ]
        )

    writer.writerow([])
    writer.writerow(["Item", "Waste (units)", "Cost", "Carbon (kg CO₂e)"])
    for entry in report["item_leaderboard"]:
        writer.writerow(
            [
                entry["item"],
                f"{entry['waste']:.2f}",
                f"{entry['cost']:.2f}",
                f"{entry['carbon']:.2f}",
            ]
        )

    writer.writerow([])
    writer.writerow(["Totals", f"{report['totals']['waste']:.2f}", f"{report['totals']['cost']:.2f}", f"{report['totals']['carbon']:.2f}"])

    csv_response = make_response(output.getvalue())
    csv_response.headers["Content-Type"] = "text/csv"
    filename = f"sustainability-event-{event_id}.csv"
    csv_response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return csv_response


@event.route(
    "/events/<int:event_id>/count_sheet/<int:location_id>",
    methods=["GET", "POST"],
)
@login_required
def count_sheet(event_id, location_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)
    el = EventLocation.query.filter_by(
        event_id=event_id, location_id=location_id
    ).first()
    if el is None:
        abort(404)
    if ev.closed:
        flash("This event is closed and cannot be modified.")
        return redirect(url_for("event.view_event", event_id=event_id))

    location, stand_items = _get_stand_items(location_id, event_id)

    if request.method == "POST":
        for entry in stand_items:
            item_id = entry["item"].id
            sheet = EventStandSheetItem.query.filter_by(
                event_location_id=el.id,
                item_id=item_id,
            ).first()
            if not sheet:
                sheet = EventStandSheetItem(
                    event_location_id=el.id, item_id=item_id
                )
                db.session.add(sheet)
            recv_qty = coerce_float(
                request.form.get(f"recv_{item_id}"), default=0.0
            ) or 0
            trans_qty = coerce_float(
                request.form.get(f"trans_{item_id}"), default=0.0
            ) or 0
            base_qty = coerce_float(
                request.form.get(f"base_{item_id}"), default=0.0
            ) or 0
            recv_factor = (
                entry["recv_unit"].factor if entry["recv_unit"] else 0
            )
            trans_factor = (
                entry["trans_unit"].factor if entry["trans_unit"] else 0
            )
            total = (
                recv_qty * recv_factor + trans_qty * trans_factor + base_qty
            )
            sheet.opening_count = recv_qty
            sheet.transferred_in = trans_qty
            sheet.transferred_out = base_qty
            sheet.closing_count = total
        el.confirmed = True
        db.session.commit()
        log_activity(
            f"Updated count sheet for event {event_id} location {location_id}"
        )
        flash("Count sheet saved")
        return redirect(url_for("event.view_event", event_id=event_id))

    return render_template(
        "events/count_sheet.html",
        event=ev,
        location=location,
        stand_items=stand_items,
    )


@event.route("/events/<int:event_id>/stand_sheets")
@login_required
def bulk_stand_sheets(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)
    data = []
    for el in ev.locations:
        loc, items = _get_stand_items(el.location_id, event_id)
        data.append(
            {
                "location": loc,
                "stand_items": items,
            }
        )
    dt = datetime.now()
    generated_at_local = (
        f"{dt.month}/{dt.day}/{dt.year} {dt.strftime('%I:%M %p').lstrip('0')}"
    )
    return render_template(
        "events/bulk_stand_sheets.html",
        event=ev,
        data=data,
        generated_at_local=generated_at_local,
    )


@event.route(
    "/events/<int:event_id>/stand_sheets/email",
    methods=["POST"],
)
@login_required
def email_bulk_stand_sheets(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)

    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    email_address = (request.form.get("email") or "").strip()
    if not email_address:
        message = "Please provide an email address."
        if is_ajax:
            return (
                jsonify({
                    "success": False,
                    "field_errors": {"email": [message]},
                    "message": message,
                }),
                400,
            )
        flash(message, "danger")
        return redirect(url_for("event.bulk_stand_sheets", event_id=event_id))

    raw_location_ids = request.form.getlist("location_ids")
    requested_location_ids = []
    saw_location_filter = False
    for raw_value in raw_location_ids:
        for token in (raw_value or "").split(","):
            token = token.strip()
            if not token:
                continue
            saw_location_filter = True
            try:
                location_id = int(token)
            except ValueError:
                message = "One or more selected locations are invalid."
                if is_ajax:
                    return jsonify({"success": False, "message": message}), 400
                flash(message, "danger")
                return redirect(url_for("event.view_event", event_id=event_id))
            if location_id not in requested_location_ids:
                requested_location_ids.append(location_id)

    event_locations_by_location_id = {
        el.location_id: el for el in ev.locations if el.location_id is not None
    }
    if saw_location_filter:
        invalid_location_ids = [
            location_id
            for location_id in requested_location_ids
            if location_id not in event_locations_by_location_id
        ]
        if invalid_location_ids:
            message = "One or more selected locations are not part of this event."
            if is_ajax:
                return jsonify({"success": False, "message": message}), 400
            flash(message, "danger")
            return redirect(url_for("event.view_event", event_id=event_id))
        event_locations = [
            event_locations_by_location_id[location_id]
            for location_id in requested_location_ids
        ]
    else:
        event_locations = list(ev.locations)

    data = []
    for el in event_locations:
        loc, items = _get_stand_items(el.location_id, event_id)
        data.append({"location": loc, "stand_items": items})

    dt = datetime.now()
    generated_at_local = (
        f"{dt.month}/{dt.day}/{dt.year} {dt.strftime('%I:%M %p').lstrip('0')}"
    )

    try:
        pdf_bytes = render_stand_sheet_pdf(
            [
                (
                    "events/bulk_stand_sheets_pdf.html",
                    {
                        "event": ev,
                        "data": data,
                        "generated_at_local": generated_at_local,
                        "pdf_export": True,
                    },
                )
            ],
            base_url=request.url_root,
        )
    except Exception:
        current_app.logger.exception(
            "Failed to render stand sheet PDF for event %s", event_id
        )
        message = "Unable to generate the stand sheet PDF."
        if is_ajax:
            return jsonify({"success": False, "message": message}), 500
        flash(message, "danger")
        return redirect(url_for("event.bulk_stand_sheets", event_id=event_id))

    try:
        stand_sheet_count = len(event_locations)
        send_email(
            to_address=email_address,
            subject=f"{ev.name} stand sheets",
            body=(
                "Attached are the stand sheets for the requested event."
                if stand_sheet_count != 1
                else "Attached is the requested stand sheet."
            ),
            attachments=[
                (
                    f"event-{event_id}-stand-sheets.pdf",
                    pdf_bytes,
                    "application/pdf",
                )
            ],
        )
    except Exception:
        current_app.logger.exception(
            "Failed to send stand sheet email for event %s", event_id
        )
        message = "Unable to send the stand sheet email."
        if is_ajax:
            return jsonify({"success": False, "message": message}), 500
        flash(message, "danger")
        return redirect(url_for("event.bulk_stand_sheets", event_id=event_id))

    log_activity(
        (
            f"Emailed stand sheet for 1 location in event {event_id} to {email_address}"
            if len(event_locations) == 1
            else f"Emailed stand sheets for event {event_id} to {email_address}"
        )
    )
    success_message = (
        f"Stand sheet sent to {email_address}."
        if len(event_locations) == 1
        else f"Stand sheets sent to {email_address}."
    )
    if is_ajax:
        return jsonify({"success": True, "message": success_message})
    flash(success_message, "success")
    return redirect(url_for("event.bulk_stand_sheets", event_id=event_id))


@event.route("/events/<int:event_id>/count_sheets")
@login_required
def bulk_count_sheets(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)
    data = []
    for el in ev.locations:
        loc, items = _get_stand_items(el.location_id, event_id)
        data.append(
            {
                "location": loc,
                "stand_items": items,
                "page_number": 1,
                "page_count": 1,
            }
        )
    return render_template(
        "events/bulk_count_sheets.html", event=ev, data=data
    )


@event.route("/events/<int:event_id>/close", methods=["POST"])
@login_required
def close_event(event_id):
    form = CSRFOnlyForm()
    if not form.validate_on_submit():
        abort(400)
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)
    if any(not el.confirmed for el in ev.locations):
        flash(
            "All locations must be confirmed before closing the event.",
            "warning",
        )
        return redirect(url_for("event.view_event", event_id=event_id))
    for el in ev.locations:
        counted_item_ids = set()
        for sheet in el.stand_sheet_items:
            counted_item_ids.add(sheet.item_id)
            lsi = LocationStandItem.query.filter_by(
                location_id=el.location_id, item_id=sheet.item_id
            ).first()
            if not sheet.closing_count:
                if lsi:
                    db.session.delete(lsi)
                continue
            if not lsi:
                lsi = LocationStandItem(
                    location_id=el.location_id,
                    item_id=sheet.item_id,
                    purchase_gl_code_id=sheet.item.purchase_gl_code_id,
                )
                db.session.add(lsi)
            elif (
                lsi.purchase_gl_code_id is None
                and sheet.item.purchase_gl_code_id is not None
            ):
                lsi.purchase_gl_code_id = sheet.item.purchase_gl_code_id
            lsi.expected_count = sheet.closing_count

        if counted_item_ids:
            LocationStandItem.query.filter(
                LocationStandItem.location_id == el.location_id,
                ~LocationStandItem.item_id.in_(counted_item_ids),
            ).delete(synchronize_session=False)
        else:
            LocationStandItem.query.filter_by(
                location_id=el.location_id
            ).delete()

        TerminalSale.query.filter_by(event_location_id=el.id).delete()

    ev.closed = True
    db.session.commit()
    log_activity(f"Closed event {event_id}")
    flash("Event closed")
    return redirect(url_for("event.view_events"))


@event.route("/events/<int:event_id>/inventory_report")
@login_required
def inventory_report(event_id):
    """Display inventory variances and GL code totals for an event."""
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)

    rows = []
    gl_totals = {}
    grand_total = 0.0

    for el in ev.locations:
        loc = el.location
        for sheet in el.stand_sheet_items:
            item = sheet.item
            lsi = LocationStandItem.query.filter_by(
                location_id=loc.id, item_id=item.id
            ).first()
            expected = lsi.expected_count if lsi else 0
            variance = sheet.closing_count - expected
            cost_total = sheet.closing_count * item.cost
            gl_obj = item.purchase_gl_code_for_location(loc.id)
            gl_code = gl_obj.code if gl_obj else "Unassigned"
            rows.append(
                {
                    "location": loc,
                    "item": item,
                    "expected": expected,
                    "actual": sheet.closing_count,
                    "variance": variance,
                    "gl_code": gl_code,
                    "cost_total": cost_total,
                }
            )
            gl_totals[gl_code] = gl_totals.get(gl_code, 0.0) + cost_total
            grand_total += cost_total

    return render_template(
        "events/inventory_report.html",
        event=ev,
        rows=rows,
        gl_totals=gl_totals,
        grand_total=grand_total,
    )
