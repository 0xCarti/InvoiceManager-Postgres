import datetime
import io
import re
import zipfile

import pytest
from openpyxl import Workbook
from werkzeug.datastructures import FileStorage

from app.models import Vendor
from app.services.purchase_imports import (
    CSVImportError,
    ParsedPurchaseLine,
    parse_purchase_order_csv,
)


def _make_pratts_file(csv_text: str) -> FileStorage:
    return FileStorage(stream=io.BytesIO(csv_text.encode()), filename="pratts.csv")


def _make_pratts_vendor() -> Vendor:
    return Vendor(first_name="Pratt", last_name="Supplies")


def _make_manitoba_vendor() -> Vendor:
    return Vendor(first_name="Manitoba", last_name="Liquor & Lotteries")


def _make_manitoba_file(*, malformed_dimension: bool = False) -> FileStorage:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Item Number",
            "Order Quantity",
            "Product Description",
            "Vol/Case Size",
            "Unit Price",
            "Extended Price",
            "Invoice No.",
            "Order No.",
            "Original Order No",
            "Invoice Date",
        ]
    )
    sheet.append(
        [
            18669,
            15,
            "COORS ORIGINAL 24/355C",
            "8520 ml x 1",
            46.57,
            698.55,
            7767932,
            2677332,
            "",
            datetime.date(2026, 4, 7),
        ]
    )
    sheet.append([None, None, "Container Deposit", None, None, 36.00])
    sheet.append(
        [
            20406,
            8,
            "OLD STYLE PILSNER 58.67L KEG",
            "58.67 L x 1",
            219.99,
            1759.92,
            7767932,
            2677332,
            "",
            datetime.date(2026, 4, 7),
        ]
    )

    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()
    payload = stream.getvalue()

    if malformed_dimension:
        source = io.BytesIO(payload)
        rewritten = io.BytesIO()
        with zipfile.ZipFile(source) as archive, zipfile.ZipFile(
            rewritten, "w"
        ) as target:
            for info in archive.infolist():
                content = archive.read(info.filename)
                if info.filename == "xl/worksheets/sheet1.xml":
                    content = re.sub(
                        br'<dimension ref="[^"]+"',
                        b'<dimension ref="A1:A1"',
                        content,
                        count=1,
                    )
                target.writestr(info, content)
        payload = rewritten.getvalue()

    return FileStorage(
        stream=io.BytesIO(payload),
        filename="manitoba_liquor.xlsx",
    )


def test_parse_pratts_csv_success():
    csv_text = """Item,Pack,Size,Brand,Description,Quantity Shipped,Unit Price,Extended Price,PO Number
1001,1,12 oz,BrandA,First Item,4,2.50,10.00,PO-123
1002,2,6 ct,BrandB,,3,1.00,3.00,PO-123
"""
    parsed = parse_purchase_order_csv(_make_pratts_file(csv_text), _make_pratts_vendor())

    assert len(parsed.items) == 2
    assert parsed.order_number == "PO-123"
    assert parsed.expected_total == 13

    first: ParsedPurchaseLine = parsed.items[0]
    assert first.vendor_sku == "1001"
    assert first.vendor_description == "First Item"
    assert first.pack_size == "1 12 oz"
    assert first.quantity == 4
    assert first.unit_cost == 2.5

    second: ParsedPurchaseLine = parsed.items[1]
    assert second.vendor_sku == "1002"
    assert second.vendor_description == "1002"
    assert second.pack_size == "2 6 ct"
    assert second.quantity == 3
    assert second.unit_cost == 1


def test_parse_pratts_csv_missing_headers():
    csv_text = """Item,Size,Brand,Description,Qty Ship,Price,Ext Price
1001,1,BrandA,First Item,4,2.50,10.00
"""

    with pytest.raises(CSVImportError) as excinfo:
        parse_purchase_order_csv(_make_pratts_file(csv_text), _make_pratts_vendor())

    assert "Missing required Pratts columns" in str(excinfo.value)
    assert "pack" in str(excinfo.value)


def test_parse_pratts_csv_invalid_quantities():
    csv_text = """Item,Pack,Size,Brand,Description,Qty Ship,Price,Ext Price
1001,1,12 oz,BrandA,First Item,0,2.50,0.00
1002,2,6 ct,BrandB,Second Item,,1.00,0.00
1003,3,5 lb,BrandC,Third Item,-1,3.00,-3.00
"""

    with pytest.raises(CSVImportError) as excinfo:
        parse_purchase_order_csv(_make_pratts_file(csv_text), _make_pratts_vendor())

    assert "No purchasable lines found" in str(excinfo.value)


def test_parse_manitoba_liquor_xlsx_success_with_malformed_dimension():
    parsed = parse_purchase_order_csv(
        _make_manitoba_file(malformed_dimension=True),
        _make_manitoba_vendor(),
    )

    assert len(parsed.items) == 2
    assert parsed.order_number == "2677332"
    assert parsed.order_date == datetime.date(2026, 4, 7)
    assert parsed.expected_total == pytest.approx(2494.47)

    first: ParsedPurchaseLine = parsed.items[0]
    assert first.vendor_sku == "18669"
    assert first.vendor_description == "COORS ORIGINAL 24/355C"
    assert first.pack_size == "8520 ml x 1"
    assert first.quantity == 15
    assert first.unit_cost == pytest.approx(46.57)

    second: ParsedPurchaseLine = parsed.items[1]
    assert second.vendor_sku == "20406"
    assert second.vendor_description == "OLD STYLE PILSNER 58.67L KEG"
    assert second.quantity == 8
    assert second.unit_cost == pytest.approx(219.99)


def test_parse_manitoba_liquor_xlsx_supports_actual_export_headers():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Item Number",
            "Order Quantity",
            "Product Description",
            "Package Size",
            "Unit Price",
            "Extended Price",
            "Distributor",
            "Case Qty",
            "Units per Case",
        ]
    )
    sheet.append(
        [
            15485,
            480,
            "COORS ORIGINAL 473C",
            "473 ml x 24",
            3.44,
            1651.20,
            "BREWERS DISTRIBUTOR LTD.",
            20,
            24,
        ]
    )
    sheet.append(
        [
            18669,
            70,
            "COORS ORIGINAL 24/355C",
            "8520 ml x 1",
            46.57,
            3259.90,
            "BREWERS DISTRIBUTOR LTD.",
            70,
            1,
        ]
    )

    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()

    parsed = parse_purchase_order_csv(
        FileStorage(stream=io.BytesIO(stream.getvalue()), filename="mbll.xlsx"),
        _make_manitoba_vendor(),
    )

    assert len(parsed.items) == 2
    assert parsed.order_number is None
    assert parsed.order_date is None
    assert parsed.expected_total == pytest.approx(4911.10)

    first: ParsedPurchaseLine = parsed.items[0]
    assert first.vendor_sku == "15485"
    assert first.pack_size == "473 ml x 24"

    second: ParsedPurchaseLine = parsed.items[1]
    assert second.vendor_sku == "18669"
    assert second.pack_size == "8520 ml x 1"


def test_parse_manitoba_liquor_xlsx_missing_headers():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Item Number", "Product Description", "Extended Price"])
    sheet.append([18669, "COORS ORIGINAL 24/355C", 698.55])

    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()

    with pytest.raises(CSVImportError) as excinfo:
        parse_purchase_order_csv(
            FileStorage(stream=io.BytesIO(stream.getvalue()), filename="bad.xlsx"),
            _make_manitoba_vendor(),
        )

    assert "Missing required Manitoba Liquor & Lotteries columns" in str(
        excinfo.value
    )
    assert "quantity" in str(excinfo.value)
